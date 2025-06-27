import datetime
import decimal
import json
import calendar
import time

import requests
from http import HTTPStatus

from django.db import transaction
from django.db.models import Min, Max, Prefetch, Subquery, Case, When, Value
from django.db import models
from django.db.models.functions import Coalesce
from django.forms import model_to_dict
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.template import loader
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from datetime import datetime, timedelta
from django.utils import timezone

from apps import accounting
from apps.accounting.api_FACT import send_guide_fact
from apps.accounting.models import Payments
from apps.accounting.sunat import send_guide, credit_note
from apps.hrm.models import Person
from apps.sales.models import Product, Brand, Family, Order, OrderDetail, Presentation, Kardex
from django.core import serializers
from apps.sales.views_sunat import ApisNetPe
from apps.user.models import User
from django.db.models import Q
import pytz
from decimal import Decimal, DivisionUndefined
import threading

APIS_TOKEN = "Bearer apis-token-3244.1KWBKUSrgYq6HNht68arg8LNsId9vVLm"
api_net = ApisNetPe(APIS_TOKEN)


class ProductList(ListView):
    model = Product
    template_name = 'sales/product_list.html'

    def get_context_data(self, **kwargs):
        # product_set = Product.objects.all().order_by('id')
        context = {
            # 'product_set': product_set
        }
        return context


@transaction.atomic
def number_order(_subsidiary=None, types=None):
    # number = Order.objects.filter(subsidiary=_subsidiary, type=types).aggregate(
    #     r=Coalesce(Max('number'), 0)).get('r')
    # return number + 1
    number = 1
    order_set = Order.objects.filter(subsidiary=_subsidiary, type=types)
    if order_set.exists():
        order_obj = order_set.last()
        number = order_obj.number
        number = number + 1
    return number


def number_order_purchase(_subsidiary=None, types=None):
    # number = Order.objects.filter(subsidiary=_subsidiary, type=types).aggregate(
    #     r=Coalesce(Max('number'), 0)).get('r')
    # return number + 1
    number = 1
    order_set = Order.objects.filter(subsidiary=_subsidiary, type=types, add='N')
    if order_set.exists():
        order_obj = order_set.last()
        number = order_obj.number
        number = number + 1
    return number


def utc_to_local(utc_dt):
    local_tz = pytz.timezone('America/Bogota')
    local_dt = utc_dt.replace(tzinfo=pytz.utc).astimezone(local_tz)
    return local_tz.normalize(local_dt)


def number_purchase(_subsidiary=None, types=None, date=None):
    year = date.year
    month = date.month
    day = date.day
    number = Order.objects.filter(subsidiary=_subsidiary, type=types, create_at__month=date.month,
                                  create_at__year=date.year).aggregate(
        r=Coalesce(Max('correlative'), 0)).get('r')
    return number + 1


number_order_lock = threading.Lock()


@csrf_exempt
def order_save(request):
    if request.method == 'POST':
        order_json = request.POST.get('order', '')
        data = json.loads(order_json)
        order_id = data['order-pk']
        document = data['document']
        state = 'R'
        date = datetime.now().date()
        current_time = datetime.now()
        types = data['type']
        person = data['person']
        license_plate = data['license_plate']
        person_obj = None
        if person:
            person_obj = Person.objects.get(id=int(person))
        total = data['total']
        total_discount = data['total_discount']
        user = request.user.id
        user_obj = User.objects.get(id=user)
        subsidiary_obj = user_obj.subsidiary
        pk = None
        parent_order = data['parent_order']
        parent_order_obj = None
        if parent_order != '0':
            parent_order_obj = Order.objects.get(id=int(parent_order))

        with number_order_lock:
            if order_id != '0' and order_id != '':
                pk = int(order_id)
                obj = Order.objects.get(id=pk)
                if obj.type == 'T':
                    correlative = number_order(subsidiary_obj, types)
                    new_order = {
                        "type": types,
                        "doc": document,
                        "number": correlative,
                        "status": state,
                        "total": decimal.Decimal(total),
                        "total_discount": decimal.Decimal(total_discount),
                        "coin": '1',
                        "change": 1,
                        "create_at": date,
                        "user": user_obj,
                        "person": person_obj,
                        "subsidiary": subsidiary_obj,
                        "license_plate": license_plate,
                        "quotation": obj.number
                    }
                    order_obj = Order.objects.create(**new_order)
                    order_obj.save()

                    if order_obj:
                        for d in data['Detail']:
                            detail = d['detail']
                            product = d['product']
                            product_obj = None
                            if product != '0' or product != '':
                                product_obj = Product.objects.get(id=int(product))
                            quantity = d['quantity']
                            q = d['q']
                            if q:
                                q = int(q)
                            else:
                                q = 0
                            unit = d['unit']
                            niu = d['niu']
                            units = decimal.Decimal(quantity)
                            if unit == 'KGM':
                                units = round(decimal.Decimal(quantity) * decimal.Decimal(niu), 4)
                            price = decimal.Decimal(d['price'])
                            is_sate = False
                            operation = None
                            invoice = False
                            if types == 'T':
                                operation = 'S'
                                is_sate = False
                                invoice = False
                            elif types == 'V':
                                operation = 'S'
                                is_sate = True
                                invoice = True
                            elif types == 'C':
                                operation = 'E'
                                is_sate = True
                                invoice = False
                            new_order_detail = {
                                "operation": operation,
                                "order": order_obj,
                                "product": product_obj,
                                "quantity": decimal.Decimal(quantity),
                                "quantity_niu": int(units),
                                "price": decimal.Decimal(round(price, 4)),
                                "unit": unit,
                                "is_state": is_sate,
                                "is_invoice": invoice
                            }
                            order_detail_obj = OrderDetail.objects.create(**new_order_detail)
                            order_detail_obj.save()
                            if order_detail_obj:
                                if order_obj.type == 'V':
                                    output_stock(order_detail_obj)
                        if order_obj:
                            if user_obj.is_authorization:
                                user_obj.is_authorization = False
                                user_obj.save()
                            return JsonResponse({
                                'success': True,
                                'order': order_obj.id,
                                'number': order_obj.number,
                                'hatch': order_obj.user.user_work,  # ventanilla
                                'userID': order_obj.user.id,  # ventanilla
                                'message': 'Proceso realizado correctamente'
                            }, status=HTTPStatus.OK)
                        else:
                            return JsonResponse({
                                'success': False,
                                'message': 'Ocurrio un problema en el proceso'
                            }, status=HTTPStatus.OK)
                else:
                    correlative = obj.number
            else:
                last_order_set = Order.objects.filter(user=user_obj)
                if last_order_set.exists():
                    last_order_obj = last_order_set.last()
                    # print("current_time", current_time.timestamp())
                    # print("last_order_obj.update_at", last_order_obj.update_at.timestamp())
                    time_difference = current_time.timestamp() - last_order_obj.update_at.timestamp()
                    # print(time_difference)
                    if time_difference >= 15:
                        correlative = number_order(subsidiary_obj, types)
                        search_correlative_set = Order.objects.filter(number=correlative)
                        if search_correlative_set.exists():
                            correlative = correlative + 1
                    else:
                        if order_id != '0' and order_id != '':
                            pk = int(order_id)
                            obj = Order.objects.get(id=pk)
                            correlative = obj.number
                        else:
                            data = {'error': "Orden Registrada"}
                            response = JsonResponse(data)
                            response.status_code = HTTPStatus.INTERNAL_SERVER_ERROR
                            return response
                else:
                    correlative = number_order(subsidiary_obj, types)
            order_obj, created = Order.objects.update_or_create(
                id=pk,
                defaults={
                    "type": types,
                    "doc": document,
                    "number": correlative,
                    "status": state,
                    "total": decimal.Decimal(total),
                    "total_discount": decimal.Decimal(total_discount),
                    "coin": '1',
                    "change": 1,
                    "create_at": date,
                    "user": user_obj,
                    "person": person_obj,
                    "subsidiary": subsidiary_obj,
                    "license_plate": license_plate,
                    "parent_order": parent_order_obj,
                })
            if order_obj:
                for d in data['Detail']:
                    detail = d['detail']
                    product = d['product']
                    product_obj = None
                    if product != '0' or product != '':
                        product_obj = Product.objects.get(id=int(product))
                    quantity = d['quantity']
                    q = d['q']
                    if q:
                        q = int(q)
                    else:
                        q = 0
                    unit = d['unit']
                    niu = d['niu']
                    units = decimal.Decimal(quantity)
                    if unit == 'KGM':
                        quantity_minimum_kgm = decimal.Decimal(get_quantity_min_kgm(product_obj))
                        units = round(decimal.Decimal(quantity) * decimal.Decimal(quantity_minimum_kgm))
                    price = decimal.Decimal(d['price'])
                    is_sate = False
                    operation = None
                    invoice = False
                    if types == 'T':
                        operation = 'S'
                        is_sate = False
                        invoice = False
                    elif types == 'V':
                        operation = 'S'
                        is_sate = True
                        invoice = True
                    elif types == 'C':
                        operation = 'E'
                        is_sate = True
                        invoice = False
                    dk = None
                    if detail != '0' and detail != '' and types != 'T':
                        dk = int(detail)
                    order_detail_obj, created_detail = OrderDetail.objects.update_or_create(
                        id=dk,
                        defaults={
                            "operation": operation,
                            "order": order_obj,
                            "product": product_obj,
                            "quantity": decimal.Decimal(quantity),
                            "quantity_niu": int(units),
                            "price": decimal.Decimal(round(price, 4)),
                            "unit": unit,
                            "is_state": is_sate,
                            "is_invoice": invoice
                        })
                    if order_detail_obj:
                        if created_detail:
                            if order_obj.type == 'V':
                                output_stock(order_detail_obj)
                            # elif order_obj.type == 'C':
                            #     input_stock(order_detail_obj)
                        else:
                            if order_obj.type == 'V':
                                output_update_stock(order_detail_obj, q)
                            # elif order_obj.type == 'C':
                            #     input_stock(order_detail_obj)
                if order_obj:
                    if user_obj.is_authorization:
                        user_obj.is_authorization = False
                        user_obj.save()
                    return JsonResponse({
                        'success': True,
                        'order': order_obj.id,
                        'number': order_obj.number,
                        'hatch': order_obj.user.user_work,  # ventanilla
                        'userID': order_obj.user.id,  # ventanilla
                        'message': 'Proceso realizado correctamente'
                    }, status=HTTPStatus.OK)
                else:
                    return JsonResponse({
                        'success': False,
                        'message': 'Ocurrio un problema en el proceso'
                    }, status=HTTPStatus.OK)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Ocurrio un problema en el proceso'
                }, status=HTTPStatus.OK)
    return JsonResponse({'message': 'Error de peticion.'}, status=HTTPStatus.BAD_REQUEST)


def output_stock(order_detail_obj=None):
    if order_detail_obj:
        product_obj = order_detail_obj.product
        quantity = order_detail_obj.quantity_niu
        product_obj.stock = product_obj.stock - decimal.Decimal(quantity)
        product_obj.save()
        order_detail_obj.quantity_remaining = decimal.Decimal(product_obj.stock)
        order_detail_obj.save()


def output_update_stock(order_detail_obj=None, q=0):
    if order_detail_obj:
        product_obj = order_detail_obj.product
        quantity = order_detail_obj.quantity_niu - q
        product_obj.stock = product_obj.stock - decimal.Decimal(quantity)
        product_obj.save()
        order_detail_obj.quantity_remaining = product_obj.stock
        order_detail_obj.save()


def input_stock(order_detail_obj=None):
    if order_detail_obj:
        product_obj = order_detail_obj.product
        quantity = order_detail_obj.quantity_niu
        product_obj.stock = product_obj.stock + quantity
        product_obj.save()
        order_detail_obj.quantity_remaining = product_obj.stock
        order_detail_obj.save()


def input_update_stock(order_detail_obj=None, q=0):
    if order_detail_obj:
        product_obj = order_detail_obj.product
        quantity = order_detail_obj.quantity_niu - q
        product_obj.stock = product_obj.stock + decimal.Decimal(quantity)
        product_obj.save()
        order_detail_obj.quantity_remaining = product_obj.stock
        order_detail_obj.save()


def modal_product(request):
    if request.method == 'GET':
        product = request.GET.get('pk', '')
        product_obj = None
        if product:
            product_obj = Product.objects.get(id=int(product))
        brand_set = Brand.objects.all()
        family_set = Family.objects.all()
        t = loader.get_template('sales/product.html')
        c = ({
            'product_obj': product_obj,
            'brand_set': brand_set,
            'family_set': family_set,
            'type_set': Product._meta.get_field('type').choices
        })
        return JsonResponse({
            'success': True,
            'form': t.render(c, request),
        })


@csrf_exempt
def product_save(request):
    if request.method == 'POST':
        try:
            product = request.POST.get('product-pk', '')
            code = request.POST.get('code', '')
            if code:
                code = code.upper()
            name = request.POST.get('name', '')
            if name:
                name = name.upper()
            description = request.POST.get('description', '')
            if description:
                description = description.upper()
            brand = request.POST.get('brand', '')
            if brand:
                brand_obj = Brand.objects.get(id=int(brand))
            else:
                brand_obj = None
            family = request.POST.get('family', '')
            if family:
                family_obj = Family.objects.get(id=int(family))
            else:
                family_obj = None
            width = request.POST.get('width', '')
            if width:
                width = width.upper()
            length = request.POST.get('length', '')
            if length:
                length = length.upper()
            height = request.POST.get('height', '')
            if height:
                height = height.upper()
            state = request.POST.get('state', False)
            is_discount = request.POST.get('is_discount', False)
            types = request.POST.get('type', 'P')
            store = request.POST.get('store', '')
            if store:
                store = store.upper()
            minimum = request.POST.get('minimum', '')
            if minimum:
                minimum = decimal.Decimal(minimum)
            else:
                minimum = decimal.Decimal(0.00)
            stock = request.POST.get('stock', '')
            if stock:
                stock = decimal.Decimal(stock)
            else:
                stock = decimal.Decimal(0.00)
            pk = None
            if state == 'on':
                state = True
            if is_discount == 'on':
                is_discount = True
            if product != '0':
                pk = int(product)
            obj, created = Product.objects.update_or_create(
                id=pk,
                defaults={
                    "code": code,
                    "name": name,
                    "description": description,
                    "brand": brand_obj,
                    "family": family_obj,
                    "width": width,
                    "length": length,
                    "height": height,
                    "is_state": state,
                    "type": types,
                    "store": store,
                    "minimum": minimum,
                    "stock": stock,
                    "is_discount": is_discount
                })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': e
            }, status=HTTPStatus.OK)
        if created:
            return JsonResponse({
                'success': True,
                'value': True,
                'product': model_to_dict(obj),
                'brand': obj.brand.name,
                'family': obj.family.name,
                'type': obj.get_type_display(),
                't': obj.type,
                'message': 'Producto registrado correctamente'
            }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': True,
                'value': False,
                'product': model_to_dict(obj),
                'brand': obj.brand.name,
                'family': obj.family.name,
                'type': obj.get_type_display(),
                't': obj.type,
                'message': 'Producto actualizado correctamente'
            }, status=HTTPStatus.OK)


def update_product(request):
    if request.method == 'GET':
        relation = request.GET.get('relation')
        pk = request.GET.get('product')
        if pk != '':
            try:
                product_obj = Product.objects.get(id=int(pk))
                product_obj.relation = relation
                product_obj.save()
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': e
                }, status=HTTPStatus.OK)
            return JsonResponse({
                'success': True,
                'message': 'Producto relacionado correctamente'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Especifique el producto'
            })


class BrandList(ListView):
    model = Brand
    template_name = 'sales/brand_list.html'


def modal_brand(request):
    if request.method == 'GET':
        brand = request.GET.get('pk', '')
        brand_obj = None
        if brand:
            brand_obj = Brand.objects.get(id=int(brand))
        t = loader.get_template('sales/brand.html')
        c = ({
            'brand_obj': brand_obj
        })
        return JsonResponse({
            'success': True,
            'form': t.render(c, request),
        })


@csrf_exempt
def brand_save(request):
    if request.method == 'POST':
        brand = request.POST.get('brand-pk', '')
        name = request.POST.get('name', '')
        pk = None
        if brand != '0':
            pk = int(brand)
        obj, created = Brand.objects.update_or_create(
            id=pk,
            defaults={
                "name": name.upper()
            })
        if created:
            return JsonResponse({
                'success': True,
                'value': True,
                'brand': model_to_dict(obj),
                'message': 'Marca registrada correctamente'
            }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': True,
                'value': False,
                'brand': model_to_dict(obj),
                'message': 'Marca actualizada correctamente'
            }, status=HTTPStatus.OK)


# def delete_brand(request):
#     if request.method == 'GET':
#         b = request.GET.get('pk', '')
#         if b:
#             brand_obj = Brand.objects.get(id=int(b))
#             brand_obj.delete()
#             return JsonResponse({
#                 'success': True,
#                 'message': 'Registro eliminado con exito',
#             }, status=HTTPStatus.OK)
#         else:
#             return JsonResponse({
#                 'success': False,
#                 'message': 'Registro incorrecto',
#             }, status=HTTPStatus.OK)


class FamilyList(ListView):
    model = Family
    template_name = 'sales/family_list.html'


def modal_family(request):
    if request.method == 'GET':
        family = request.GET.get('pk', '')
        family_obj = None
        if family:
            family_obj = Family.objects.get(id=int(family))
        t = loader.get_template('sales/family.html')
        c = ({
            'family_obj': family_obj
        })
        return JsonResponse({
            'success': True,
            'form': t.render(c, request),
        })


@csrf_exempt
def family_save(request):
    if request.method == 'POST':
        family = request.POST.get('family-pk', '')
        name = request.POST.get('name', '')
        pk = None
        if family != '0':
            pk = int(family)
        obj, created = Family.objects.update_or_create(
            id=pk,
            defaults={
                "name": name.upper()
            })
        if created:
            return JsonResponse({
                'success': True,
                'value': True,
                'family': model_to_dict(obj),
                'message': 'Familia registrada correctamente'
            }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': True,
                'value': False,
                'family': model_to_dict(obj),
                'message': 'Familia actualizada correctamente'
            }, status=HTTPStatus.OK)


def modal_presenting(request):
    if request.method == 'GET':
        product = request.GET.get('pk', '')
        product_obj = None
        if product:
            product_obj = Product.objects.get(id=int(product))
            t = loader.get_template('sales/presentation.html')
            c = ({
                'product_obj': product_obj,
                'unit_set': Presentation._meta.get_field('unit').choices
            })
            return JsonResponse({
                'success': True,
                'form': t.render(c, request),
            })


@csrf_exempt
def presentation_save(request):
    if request.method == 'POST':
        product = request.POST.get('product-pk', '')
        price = request.POST.get('price', 0)
        unit = request.POST.get('unit', 0)
        corporate_sale = request.POST.get('corporate_sale', 0)
        _corporate_sale = False
        if price:
            price = decimal.Decimal(price)
        else:
            price = decimal.Decimal(0.00)
        quantity = request.POST.get('quantity', 0)
        if quantity:
            quantity = decimal.Decimal(quantity)
        else:
            quantity = decimal.Decimal(0.00)
        quantity_niu = request.POST.get('quantity_niu', 0)
        if quantity_niu:
            quantity_niu = int(quantity_niu)
        else:
            if unit == 'NIU':
                quantity_niu = int(quantity)
            elif unit == 'KGM':
                quantity_niu = 1
            else:
                quantity_niu = 1
        if corporate_sale != 0:
            _corporate_sale = True

        if product != '':
            product_obj = Product.objects.get(id=int(product))
            presentation_obj = Presentation(
                product=product_obj,
                unit=unit,
                price=price,
                quantity=quantity,
                quantity_niu=quantity_niu,
                is_corporate=_corporate_sale
            )
            presentation_obj.save()
            return JsonResponse({
                'success': True,
                'id': presentation_obj.id,
                'price': presentation_obj.price,
                'quantity': presentation_obj.quantity,
                'unit': presentation_obj.get_unit_display(),
                'quantity_niu': presentation_obj.quantity_niu,
                'is_corporate': presentation_obj.is_corporate,
                'message': 'Presentación registrada correctamente'
            }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'Especifique un producto'
            }, status=HTTPStatus.OK)


def delete_presentation(request):
    if request.method == 'GET':
        p = request.GET.get('pk', '')
        if p:
            presentation_obj = Presentation.objects.get(id=int(p))
            presentation_obj.delete()
            return JsonResponse({
                'success': True,
                'message': 'Registro eliminado con exito',
            }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'Registro incorrecto',
            }, status=HTTPStatus.OK)


def get_store(request):
    if request.method == 'GET':
        product = request.GET.get('pk', '')
        product_obj = None
        if product:
            product_obj = Product.objects.get(id=int(product))
            # subsidiary_store_set = SubsidiaryStore.objects.all()
            # store_set = Store.objects.filter(product=product_obj)
            t = loader.get_template('sales/inventory.html')
            c = ({
                'product_obj': product_obj,
                # 'subsidiary_store_set': subsidiary_store_set,
                # 'store_set': store_set
            })
            return JsonResponse({
                'success': True,
                'form': t.render(c, request),
            })


@csrf_exempt
def store_save(request):
    if request.method == 'POST':
        store = request.POST.get('store-pk', '')
        product = request.POST.get('product-pk', '')
        stock = request.POST.get('stock', 0)
        minimum = request.POST.get('minimum', 0)
        subsidiary_store = request.POST.get('subsidiary_store', '')
        if product != '' and subsidiary_store != '':
            product_obj = Product.objects.get(id=int(product))
            # subsidiary_store_obj = SubsidiaryStore.objects.get(id=int(subsidiary_store))
            pk = None
            if store != '0' and store != '':
                pk = int(store)
            # obj, created = Store.objects.update_or_create(
            #     id=pk,
            #     defaults={
            #         "product": product_obj,
            #         "subsidiary_store": subsidiary_store_obj,
            #         "stock": decimal.Decimal(stock),
            #         "minimum": decimal.Decimal(minimum)
            #     })
            # return JsonResponse({
            #     'success': True,
            #     'id': obj.id,
            #     'store': obj.subsidiary_store.name,
            #     's': obj.subsidiary_store.id,
            #     'stock': obj.stock,
            #     'minimum': obj.minimum,
            #     'message': 'Proceso realizado correctamente'
            # }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'Especifique un producto o almacén'
            }, status=HTTPStatus.OK)


class SalesList(ListView):
    model = Product
    template_name = 'sales/order.html'

    def get_context_data(self, **kwargs):
        context = {
            'document_set': Person._meta.get_field('document').choices,
            'order_type_set': Order._meta.get_field('type').choices,
            'doc_set': Order._meta.get_field('doc').choices,
        }
        return context


def search_product(request):
    if request.method == 'GET':
        search = request.GET.get('search')
        product = []
        if search:
            product_set = Product.objects.filter(name__icontains=search)
            for p in product_set:
                product.append({
                    'pk': p.id,
                    'code': p.code,
                    'name': p.name,
                    'measure': str(p.width) + "x" + str(p.length) + "x" + str(p.height),
                    'stock': p.stock,
                })
        return JsonResponse({
            'status': True,
            'product': product
        })


def delete_order_detail(request):
    if request.method == 'GET':
        pk = request.GET.get('pk', '')
        sum = 0
        if pk:
            order_detail_obj = OrderDetail.objects.get(id=int(pk))
            order_obj = order_detail_obj.order
            order_detail_obj.is_state = False
            order_detail_obj.save()
            if order_obj.type == 'T':
                order_detail_obj.delete()
            if order_obj.type == 'V':
                new_order_detail = {
                    "operation": 'E',
                    "order": order_obj,
                    "product": order_detail_obj.product,
                    "quantity": decimal.Decimal(order_detail_obj.quantity),
                    "quantity_niu": int(order_detail_obj.quantity_niu),
                    "price": decimal.Decimal(round(order_detail_obj.price, 6)),
                    "unit": order_detail_obj.unit,
                    "is_state": False,
                    "is_invoice": False
                }
                new_order_detail_obj = OrderDetail.objects.create(**new_order_detail)
                new_order_detail_obj.save()
                input_stock(new_order_detail_obj)
            elif order_obj.type == 'C':
                new_order_detail = {
                    "operation": 'S',
                    "order": order_obj,
                    "product": order_detail_obj.product,
                    "quantity": decimal.Decimal(order_detail_obj.quantity),
                    "quantity_niu": int(order_detail_obj.quantity_niu),
                    "price": decimal.Decimal(round(order_detail_obj.price, 6)),
                    "unit": order_detail_obj.unit,
                    "is_state": False,
                    "is_invoice": False
                }
                new_order_detail_obj = OrderDetail.objects.create(**new_order_detail)
                new_order_detail_obj.save()
                output_stock(new_order_detail_obj)
            # order_detail_obj.is_considered = False
            # order_detail_obj.save()
            for s in order_obj.orderdetail_set.filter(is_state=True):
                sum += s.price * s.quantity
            order_obj.total = decimal.Decimal(sum)
            order_obj.save()
            # payment_set = Payments.objects.filter(order=order_obj)
            # payment_set.delete()

            return JsonResponse({
                'success': True,
                'message': 'Detalle eliminado correctamente',
            }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'No se logro indetificar el detalle',
            }, status=HTTPStatus.OK)


def update_order_detail(request):
    if request.method == 'GET':
        od = request.GET.get('od')
        value = request.GET.get('value')
        if od != '0' and od != '':
            order_detail_obj = OrderDetail.objects.get(id=int(od))
            order_detail_obj.is_invoice = bool(int(value))
            order_detail_obj.save()
            if order_detail_obj.is_invoice:
                message = 'Seleccionado'
            else:
                message = 'Deseleccionado'
            return JsonResponse({
                'status': True,
                'message': message
            })
        else:
            return JsonResponse({
                'status': False,
                'message': 'Especifique el detalle'
            })


def get_quantity_min_kgm(product):
    quantity_minimum = ''
    presenting_set = Presentation.objects.filter(product=product, unit='KGM', quantity=decimal.Decimal(1.0000))
    if presenting_set.exists():
        presenting_obj = presenting_set.first()
        quantity_minimum = presenting_obj.quantity_niu
    return quantity_minimum


prices_lock = threading.Lock()


def get_prices(request):
    if request.method == 'GET':
        quantity = request.GET.get('quantity', '')
        # print(quantity)
        product = request.GET.get('product', '')
        y = request.GET.get('y', '')
        q = request.GET.get('q', '')
        pk = request.GET.get('pk', '')
        is_corporate = request.GET.get('is_corporate', '')
        client = request.GET.get('client', '')
        state = bool(int(request.GET.get('state', 0)))
        
        # Obtener el descuento del cliente si se proporciona
        client_discount = 0
        if client and client != '' and client != '0':
            try:
                person_obj = Person.objects.get(id=int(client))
                if person_obj.discount:
                    client_discount = float(person_obj.discount.value)
            except Person.DoesNotExist:
                pass
        
        with prices_lock:
            if quantity and product:
                product_obj = Product.objects.get(id=int(product))
                if product_obj.stock:
                    stock = product_obj.stock
                else:
                    stock = decimal.Decimal(0.00)
                if str(quantity).find(".") > 0 or str(quantity).find(",") > 0:

                    if is_corporate:
                        if decimal.Decimal(quantity) <= decimal.Decimal(1.0000):
                            query_set = Presentation.objects.filter(product=product_obj,
                                                                    quantity=decimal.Decimal(quantity),
                                                                    unit='KGM', is_corporate=True).select_related(
                                'product')
                            if query_set.exists():
                                presenting_set = query_set
                            else:
                                presenting_set = Presentation.objects.filter(product=product_obj, quantity=1,
                                                                             unit='KGM',
                                                                             is_corporate=True).select_related(
                                    'product')
                        else:
                            presenting_set = Presentation.objects.filter(product=product_obj,
                                                                         quantity__lte=quantity,
                                                                         unit='KGM',
                                                                         is_corporate=True).select_related(
                                'product').order_by('quantity')
                    else:
                        if decimal.Decimal(quantity) <= decimal.Decimal(1.0000):
                            query_set = Presentation.objects.filter(product=product_obj,
                                                                    quantity=decimal.Decimal(quantity),
                                                                    unit='KGM', is_corporate=False).select_related(
                                'product')
                            if query_set.exists():
                                presenting_set = query_set
                            else:
                                presenting_set = Presentation.objects.filter(product=product_obj, quantity=1,
                                                                             unit='KGM',
                                                                             is_corporate=False).select_related(
                                    'product')
                        else:
                            presenting_set = Presentation.objects.filter(product=product_obj, quantity__lte=quantity,
                                                                         unit='KGM',
                                                                         is_corporate=False).select_related(
                                'product').order_by(
                                'quantity')
                    if presenting_set.exists():
                        presenting_obj = presenting_set.last()
                        price = decimal.Decimal(presenting_obj.price)
                        
                        # Aplicar descuento del cliente si existe
                        if client_discount > 0:
                            price = price * (1 - decimal.Decimal(client_discount / 100))
                        
                        if pk != "0" and pk != "":
                            if decimal.Decimal(quantity) <= decimal.Decimal(q):
                                return JsonResponse({
                                    'success': True,
                                    'price': round(price, 4),
                                    'niu': presenting_obj.quantity_niu,
                                    'unit': presenting_obj.unit,
                                    'amount': round(price, 4) * (round(decimal.Decimal(quantity), 4)),
                                    'message': 'Precio S/. ' + str(round(price, 4)),
                                }, status=HTTPStatus.OK)
                            else:
                                if (decimal.Decimal(quantity) - decimal.Decimal(
                                        q)) * presenting_obj.quantity_niu <= stock:
                                    return JsonResponse({
                                        'success': True,
                                        'price': round(price, 4),
                                        'niu': presenting_obj.quantity_niu,
                                        'unit': presenting_obj.unit,
                                        'amount': round(price, 4) * (round(decimal.Decimal(quantity), 4)),
                                        'message': 'Precio S/. ' + str(round(price, 4)),
                                    }, status=HTTPStatus.OK)
                                else:
                                    return JsonResponse({
                                        'success': False,
                                        'message': 'Stock Insuficiente',
                                    }, status=HTTPStatus.OK)
                        else:
                            quantity_minimum_kgm = decimal.Decimal(get_quantity_min_kgm(product_obj))
                            if decimal.Decimal(quantity) * quantity_minimum_kgm <= stock:
                                return JsonResponse({
                                    'success': True,
                                    'price': round(price, 4),
                                    'niu': decimal.Decimal(quantity) * quantity_minimum_kgm,
                                    'unit': presenting_obj.unit,
                                    'amount': round(price, 4) * (round(decimal.Decimal(quantity), 4)),
                                    'message': 'Precio S/. ' + str(round(price, 4)),
                                }, status=HTTPStatus.OK)
                            else:
                                return JsonResponse({
                                    'success': False,
                                    'message': 'Stock Insuficiente',
                                }, status=HTTPStatus.OK)
                    else:
                        return JsonResponse({
                            'success': False,
                            'message': 'No existe precios en KILOS',
                        }, status=HTTPStatus.OK)
                else:
                    user_id = request.user.id
                    user_obj = User.objects.get(id=int(user_id))
                    if pk != "0" and pk != "":
                        if decimal.Decimal(quantity) <= decimal.Decimal(q):
                            presenting_set = []
                            if user_obj.is_authorization and state:
                                presenting_set = Presentation.objects.filter(
                                    product=product_obj, quantity__gt=quantity, is_corporate=False,
                                    unit='NIU').order_by('quantity').select_related('product')
                            else:
                                presenting_set = Presentation.objects.filter(
                                    product=product_obj, quantity__lte=quantity, is_corporate=False,
                                    unit='NIU').order_by('quantity').select_related('product')
                            if presenting_set.exists():
                                presenting_obj = None
                                if user_obj.is_authorization and state:
                                    presenting_obj = presenting_set.first()
                                else:
                                    presenting_obj = presenting_set.last()
                                price = decimal.Decimal(presenting_obj.price)
                                if product_obj.is_discount and int(y):
                                    price = price * decimal.Decimal(0.5000)
                                
                                # Aplicar descuento del cliente si existe
                                if client_discount > 0:
                                    price = price * (1 - decimal.Decimal(client_discount / 100))

                                return JsonResponse({
                                    'success': True,
                                    'price': round(price, 4),
                                    'niu': presenting_obj.quantity_niu,
                                    'unit': presenting_obj.unit,
                                    'amount': round(price, 4) * (round(decimal.Decimal(quantity), 4)),
                                    'message': 'Precio S/. ' + str(round(price, 4)),
                                }, status=HTTPStatus.OK)
                        else:
                            if (decimal.Decimal(quantity) - decimal.Decimal(q)) <= stock:
                                presenting_set = Presentation.objects.filter(product=product_obj,
                                                                             quantity__lte=quantity,
                                                                             is_corporate=False,
                                                                             unit='NIU').order_by(
                                    'quantity').select_related('product')
                                if presenting_set.exists():
                                    presenting_obj = presenting_set.last()
                                    price = decimal.Decimal(presenting_obj.price)
                                    if product_obj.is_discount and int(y):
                                        price = price * decimal.Decimal(0.5000)
                                    
                                    # Aplicar descuento del cliente si existe
                                    if client_discount > 0:
                                        price = price * (1 - decimal.Decimal(client_discount / 100))
                                    
                                    return JsonResponse({
                                        'success': True,
                                        'price': round(price, 4),
                                        'niu': presenting_obj.quantity_niu,
                                        'unit': presenting_obj.unit,
                                        'amount': round(price, 4) * (round(decimal.Decimal(quantity), 4)),
                                        'message': 'Precio S/. ' + str(round(price, 4)),
                                    }, status=HTTPStatus.OK)
                                else:
                                    return JsonResponse({
                                        'success': False,
                                        'message': 'No existe precios en UNIDADES',
                                    }, status=HTTPStatus.OK)
                            else:
                                return JsonResponse({
                                    'success': False,
                                    'message': 'Stock Insuficiente',
                                }, status=HTTPStatus.OK)
                    else:
                        if decimal.Decimal(quantity) <= stock:
                            presenting_set = []
                            if user_obj.is_authorization and state:
                                presenting_set = Presentation.objects.filter(product=product_obj, quantity__gt=quantity,
                                                                             is_corporate=False,
                                                                             unit='NIU').order_by(
                                    'quantity').select_related('product')
                            else:
                                if is_corporate:
                                    presenting_set = Presentation.objects.filter(
                                        product=product_obj, quantity__lte=quantity, is_corporate=True,
                                        unit='NIU').order_by('quantity').select_related('product')
                                else:
                                    presenting_set = Presentation.objects.filter(
                                        product=product_obj,
                                        quantity__lte=quantity,
                                        is_corporate=False,
                                        unit='NIU'
                                    ).annotate(
                                        max_quantity=Max('quantity')
                                    ).order_by('-max_quantity')
                            if presenting_set.exists():
                                presenting_obj = None
                                if user_obj.is_authorization and state:
                                    presenting_obj = presenting_set.first()
                                else:
                                    presenting_obj = presenting_set.first()
                                price = decimal.Decimal(presenting_obj.price)
                                if product_obj.is_discount and int(y):
                                    price = price * decimal.Decimal(0.5000)
                                
                                # Aplicar descuento del cliente si existe
                                if client_discount > 0:
                                    price = price * (1 - decimal.Decimal(client_discount / 100))
                                
                                return JsonResponse({
                                    'success': True,
                                    'price': round(price, 4),
                                    'niu': presenting_obj.quantity_niu,
                                    'unit': presenting_obj.unit,
                                    'amount': round(price, 4) * (round(decimal.Decimal(quantity), 4)),
                                    'message': 'Precio S/. ' + str(round(price, 4)),
                                }, status=HTTPStatus.OK)
                            else:
                                return JsonResponse({
                                    'success': False,
                                    'message': 'No existe precios en UNIDADES',
                                }, status=HTTPStatus.OK)
                        else:
                            return JsonResponse({
                                'success': False,
                                'message': 'Stock Insuficiente',
                            }, status=HTTPStatus.OK)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Ingrese una cantidad',
                }, status=HTTPStatus.OK)


def get_mayor(product_obj=None):
    if product_obj:
        presenting_set = Presentation.objects.filter(product=product_obj, is_elderly=True)
        if presenting_set.exists():
            quantity = presenting_set.first().quantity_niu
            return quantity
        else:
            return 0


def get_presentation_corporate_sale(product_obj=None):
    if product_obj:
        presenting_set = Presentation.objects.filter(product=product_obj, is_corporate=True)
        if presenting_set.exists():
            quantity = presenting_set.first().quantity_niu
            return quantity
        else:
            return 0


def get_products(request):
    if request.method == 'GET':
        code = request.GET.get('code', '')
        is_corporate = request.GET.get('is_corporate', '')
        if code:
            p_set = Product.objects.filter(code=code)
            product_dic = []
            if p_set.exists():
                product = p_set.first()
                m = get_mayor(product)
                if is_corporate:
                    m = get_presentation_corporate_sale(product)
                product_dic.append(
                    {
                        'id': product.id,
                        'code': product.code,
                        'name': product.name,
                        'measure': product.measure(),
                        'm': m,
                        'y': 0,
                        'stock': product.stock,
                        'min': product.minimum
                    }
                )
                if product.relation:
                    codes = product.relation.strip().split('-')
                    if len(codes) > 0:
                        for c in codes:
                            product_set = Product.objects.filter(code=c.strip())
                            if product_set.exists():
                                p = product_set.first()
                                product_dic.append(
                                    {
                                        'id': p.id,
                                        'code': p.code,
                                        'name': p.name,
                                        'measure': p.measure(),
                                        'm': get_mayor(p),
                                        'y': 1,
                                        'stock': p.stock
                                    }
                                )
                    else:
                        return JsonResponse({
                            'success': False,
                            'message': 'El codigo no se encuentra registrado'
                        }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'El codigo no se encuentra registrado'
            }, status=HTTPStatus.OK)

        return JsonResponse({
            'success': True,
            'product': product_dic
        }, status=HTTPStatus.OK)
    else:
        return JsonResponse({
            'success': False,
            'message': 'El codigo no se encuentra registrado'
        }, status=HTTPStatus.OK)


def query_api_facturacioncloud(nro_doc):
    context = {}
    url = {}
    # type_document == 'DNI'
    url = 'http://www.facturaelectronicape.com/facturacion/controller/ws_consulta_rucdni_v2.php?usuario' \
          '=20498189637&password=marvisur.123.&documento=DNI&nro_documento=' + nro_doc
    # elif type_document == 'RUC':
    #     url = 'http://www.facturaelectronicape.com/facturacion/controller/ws_consulta_rucdni_v2.php?usuario' \
    #           '=20498189637&password=marvisur.123.&documento=RUC&nro_documento=' + nro_doc

    headers = {
        "Content-Type": 'application/json'
    }

    try:
        response = requests.post(url, headers=headers, timeout=2.5)

        if response.status_code == 200:
            result = response.json()

            context = {
                'success': result.get("success"),
                'statusMessage': result.get("statusMessage"),
                'result': result.get('result'),
                'DNI': result.get('result').get('DNI'),
                'Nombre': result.get('result').get('Nombre'),
                'Paterno': result.get('result').get('Paterno'),
                'Materno': result.get('result').get('Materno'),
                'ruc': result.get('result').get('RUC'),
                'razonSocial': result.get('result').get('RazonSocial'),
                'direccion': result.get('result').get('Direccion'),
                'Estado': result.get('result').get('Estado'),
            }
        else:
            result = response.status_code
            context = {
                'errors': True
            }
        return context

    except requests.ReadTimeout:
        print("READ TIME OUT FACTURACION CLOUD")

        context = {
            'errors': True
        }
    return context


def get_person_by_document(request):
    if request.method == 'GET':
        number = request.GET.get('number', '')
        document = request.GET.get('document', '')
        order = request.GET.get('order', '')
        t = request.GET.get('type', '')
        order_obj = None
        try:
            person_search = Person.objects.get(number=number)
        except Person.DoesNotExist:
            person_search = None
        if person_search is not None:
            if order:
                order_obj = Order.objects.get(id=int(order))
                order_obj.person = person_search
                order_obj.save()
            
            # Obtener el descuento del cliente
            discount_value = None
            if person_search.discount:
                discount_value = float(person_search.discount.value)
            
            return JsonResponse({
                'success': True,
                'id': person_search.id,
                'names': person_search.names,
                'phone': person_search.phone,
                'email': person_search.email,
                'address': person_search.address,
                'discount': discount_value,
                'message': 'Proceso exitoso'},
                status=HTTPStatus.OK)
        else:
            person_obj = None
            if document == '1' and number != '1':
                r = api_net.get_person(number)
                if r.get('success') is True:
                    name = r.get('nombres')
                    paternal = r.get('apellidoPaterno')
                    maternal = r.get('apellidoMaterno')
                    address = r.get('direccion')
                    names = str(name + ' ' + paternal + ' ' + maternal).upper()
                    if names is not None and len(names) > 0:
                        person_obj = Person(
                            document=document,
                            number=number,
                            names=names,
                            address=address,
                            type=t
                        )
                        person_obj.save()
                    else:
                        return JsonResponse({
                            'success': False,
                            'message': 'No se encontro resultados'
                        })
                # else:
                #     return JsonResponse({
                #         'success': False,
                #         'message': 'No se encontro resultados'
                #     })
                else:
                    r = query_api_facturacioncloud(number)

                    name = r.get('Nombre')
                    paternal_name = r.get('Paterno')
                    maternal_name = r.get('Materno')

                    if name is not None and r.get('statusMessage') != 'SERVICIO SE VENCIO' and 'errors' not in r:
                        if paternal_name is not None and len(paternal_name) > 0:
                            result = name + ' ' + paternal_name + ' ' + maternal_name
                            if len(result.strip()) != 0:
                                person_obj = Person(
                                    document=document,
                                    number=number,
                                    names=result,
                                    # address=address,
                                    type=t
                                )
                                person_obj.save()
                            else:
                                data = {'error': 'NO EXISTE DNI. REGISTRE MANUALMENTE'}
                                response = JsonResponse(data)
                                response.status_code = HTTPStatus.INTERNAL_SERVER_ERROR
                                return response
                    else:
                        return JsonResponse({
                            'success': False,
                            'message': 'No se encontro resultados'
                        })

            elif document == '6' and number != '1':
                r = api_net.get_company(number)
                if r.get('success') is True:
                    names = r.get('nombre').upper()
                    address = (r.get('direccion')).strip()
                    person_obj = Person(
                        document=document,
                        number=number,
                        names=names,
                        address=address,
                        type=t
                    )
                    person_obj.save()
                else:
                    return JsonResponse({
                        'success': False,
                        'message': 'No se encontro resultados'
                    })
            elif document == '1' and number == '1':
                person_obj = Person(
                    document=document,
                    number=number,
                    names='',
                    address='',
                    type=t
                )
                person_obj.save()
            if order:
                order_obj = Order.objects.get(id=int(order))
                order_obj.person = person_obj
                order_obj.save()
            
            # Obtener el descuento del cliente (nuevo o existente)
            discount_value = None
            if person_obj and person_obj.discount:
                discount_value = float(person_obj.discount.value)
            
            return JsonResponse({
                'success': True,
                'id': person_obj.id,
                'names': person_obj.names,
                'phone': person_obj.phone,
                'email': person_obj.email,
                'address': person_obj.address,
                'discount': discount_value,
                'message': 'Proceso exitoso'},
                status=HTTPStatus.OK)


def get_order(request):
    if request.method == 'GET':
        number = request.GET.get('number', '')
        order_type = request.GET.get('type', '')
        type_name = request.GET.get('type_name', '')
        user_id = request.user.id
        user_obj = User.objects.get(id=user_id)
        subsidiary_obj = user_obj.subsidiary
        order_set = Order.objects.filter(type=order_type, subsidiary=subsidiary_obj, number=number)
        if order_set.exists():
            order_obj = order_set.last()
            if order_obj.condition not in ['PA', 'A']:
                if order_obj.status == 'R':
                    payment = '0'
                    payment_set = Payments.objects.filter(order=order_obj)
                    if payment_set.exists():
                        payment = payment_set.last().id
                    detail = []
                    if order_type == 'T':
                        order_detail_set = OrderDetail.objects.filter(order=order_obj)
                    else:
                        order_detail_set = OrderDetail.objects.filter(order=order_obj, is_state=True)
                    person = ''
                    person_number = ''
                    person_names = ''
                    person_address = ''
                    relation = None
                    for d in order_detail_set.order_by('id'):
                        y = 0
                        if d.product.relation != '' and d.product.relation != '-':
                            relation = d.product.relation
                        if relation == d.product.code:
                            y = 1
                            relation = None
                        row = {
                            'id': d.id,
                            'code': d.product.code,
                            'product': d.product.id,
                            'description': d.product.name,
                            'measure': d.product.measure(),
                            'unit': d.unit,
                            'quantity': d.quantity,
                            'quantity_niu': d.quantity_niu,
                            'price': d.price,
                            'm': get_mayor(d.product),
                            'y': y,
                            'order_id': d.order.id,
                            'type_order': d.order.type,
                            'stock': d.product.stock
                        }
                        detail.append(row)

                        if order_obj.person:
                            person = order_obj.person.id
                            person_number = order_obj.person.number,
                            person_names = order_obj.person.names,
                            person_address = order_obj.person.address,
                    return JsonResponse({
                        'success': True,
                        'pk': order_obj.id,
                        'last_order': order_obj.id,
                        'type': order_obj.type,
                        'discount': order_obj.total_discount,
                        'license_plate': order_obj.license_plate,
                        'doc': order_obj.doc,
                        'person': person,
                        'person_number': person_number,
                        'person_names': person_names,
                        'person_address': person_address,
                        'detail': detail,
                        'payment': payment,
                    }, status=HTTPStatus.OK)
                else:
                    return JsonResponse({
                        'success': False,
                        'message': 'Orden ' + str(order_obj.get_status_display()).lower(),
                    }, status=HTTPStatus.OK)
            else:
                return JsonResponse({
                    'success': False,
                    'message': '¡Orden Anulada!',
                }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'No existe la Orden de ' + str(type_name).capitalize() + ' Nº ' + str(number),
            }, status=HTTPStatus.OK)


class PurchaseList(ListView):
    model = Product
    template_name = 'sales/purchase.html'

    def get_context_data(self, **kwargs):
        my_date = datetime.now()
        date_now = my_date.strftime("%Y-%m-%d")
        context = {
            'document_set': Person._meta.get_field('document').choices,
            'unit_set': Presentation._meta.get_field('unit').choices,
            'date_now': date_now,
            'coin_set': Order._meta.get_field('coin').choices,
            'type_set': Order._meta.get_field('doc').choices
        }
        return context


def sales_list(request):
    if request.method == 'GET':
        sales_set = Order.objects.filter(type='V').order_by('id')
        return render(request, 'sales/sales_list.html', {
            'sales_set': sales_set
        })


def modal_guide(request):
    if request.method == 'GET':
        pk = request.GET.get('pk', '')
        if pk:
            my_date = datetime.now()
            date_now = my_date.strftime("%Y-%m-%d")
            order_obj = Order.objects.get(id=int(pk))
            motive_set = Order._meta.get_field('guide_motive').choices
            guide_modality_set = Order._meta.get_field('guide_modality_transport').choices
            tpl = loader.get_template('sales/guide.html')
            context = ({
                'date_now': date_now,
                'order_obj': order_obj,
                'guide_motive_set': motive_set,
                'guide_modality_set': guide_modality_set
            })
            return JsonResponse({
                'grid': tpl.render(context, request),
            }, status=HTTPStatus.OK, content_type="application/json")
        else:
            return JsonResponse({
                'success': False,
                'message': 'Orden desconocida',
            }, status=HTTPStatus.OK)


@csrf_exempt
def guide_save(request):
    if request.method == 'POST':
        order = request.POST.get('id-order', '')
        if order:
            order_obj = Order.objects.get(id=int(order))
            description = ''
            if order_obj.doc == '1':
                description = 'FACTURA ELECTRONICA: ' + str(order_obj.bill_serial) + '-' + str(order_obj.bill_number)
            elif order_obj.doc == '2':
                description = 'BOLETA ELECTRONICA: ' + str(order_obj.bill_serial) + '-' + str(order_obj.bill_number)
            date_issue = request.POST.get('issue', '')
            date_transfer = request.POST.get('transfer', '')

            motive = request.POST.get('motive', '')
            truck = request.POST.get('truck', '')

            driver_name = request.POST.get('driver-names', '')
            driver_last_name = request.POST.get('driver-lastname', '')

            driver_dni = request.POST.get('driver-dni', '')
            driver_license = request.POST.get('driver-license', '')

            guide_carrier_document = request.POST.get('guide_carrier_document', '')
            guide_carrier_names = request.POST.get('guide_carrier_names', '')

            guide_origin_address = request.POST.get('subsidiary-address', '')
            guide_destiny_address = request.POST.get('person-address', '')

            origin = request.POST.get('origin', '')
            modality_transport = request.POST.get('modality_transport', '')
            destiny = request.POST.get('destiny', '')
            weight = request.POST.get('weight', 0)
            lumps = request.POST.get('lumps', 0)
            register_mtc = request.POST.get('register_mtc', '')

            # if request.POST.get('observation', '') != '':
            description = str(request.POST.get('observation', ''))

            subsidiary_obj = order_obj.subsidiary
            serial = 'T0' + str(subsidiary_obj.serial)
            correlative = accounting.views.number_guide(subsidiary=subsidiary_obj, document='G')
            order_obj.guide_serial = serial
            order_obj.add = 'G'
            order_obj.guide_description = description
            order_obj.guide_type = 7
            order_obj.guide_number = correlative
            order_obj.guide_date = date_issue
            order_obj.guide_transfer = date_transfer
            order_obj.guide_motive = motive
            order_obj.guide_truck = truck
            order_obj.guide_driver_name = driver_name
            order_obj.guide_driver_lastname = driver_last_name
            order_obj.guide_driver_full_name = driver_name + ' ' + driver_last_name
            order_obj.guide_driver_dni = driver_dni
            order_obj.guide_driver_license = driver_license
            order_obj.guide_origin = origin
            order_obj.guide_destiny = destiny
            order_obj.guide_carrier_document = guide_carrier_document
            order_obj.guide_modality_transport = modality_transport
            order_obj.guide_carrier_names = guide_carrier_names
            order_obj.guide_register_mtc = register_mtc

            order_obj.guide_origin_address = guide_origin_address
            order_obj.guide_destiny_address = guide_destiny_address

            if weight:
                order_obj.guide_weight = decimal.Decimal(weight)
            if lumps:
                order_obj.guide_package = decimal.Decimal(lumps)
            order_obj.save()

            if order_obj.id:
                # r = send_guide(order_obj.id)
                r = send_guide_fact(order_obj.id)
                if r.get('success'):
                    return JsonResponse({
                        'success': True,
                        'message': r.get('message'),
                        'pk': order_obj.id
                    }, status=HTTPStatus.OK)
                else:
                    order_obj.guide_serial = '-'
                    order_obj.add = 'N'
                    order_obj.guide_description = '-'
                    order_obj.guide_type = ''
                    order_obj.guide_number = None
                    order_obj.guide_date = None
                    order_obj.guide_transfer = None
                    order_obj.guide_motive = '01'
                    order_obj.guide_truck = ''
                    order_obj.guide_driver_name = ''
                    order_obj.guide_driver_lastname = ''
                    order_obj.guide_driver_full_name = ''
                    order_obj.guide_driver_dni = ''
                    order_obj.guide_driver_license = ''
                    order_obj.guide_origin = ''
                    order_obj.guide_destiny = ''
                    order_obj.guide_carrier_document = ''
                    order_obj.guide_modality_transport = '1'
                    order_obj.guide_carrier_names = ''
                    order_obj.guide_origin_address = ''
                    order_obj.guide_destiny_address = ''
                    order_obj.guide_weight = decimal.Decimal(0.0000)
                    order_obj.guide_package = decimal.Decimal(0.0000)
                    order_obj.guide_register_mtc = ''
                    order_obj.save()
                    return JsonResponse({
                        'success': False,
                        'message': r.get('errors'),
                    }, status=HTTPStatus.OK)

            # return JsonResponse({
            #     'success': True,
            #     'message': 'Guia realizada con exito',
            #     'pk': order_obj.id
            # }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'Orden no idetificada'
            }, status=HTTPStatus.OK)


def get_product_purchase(request):
    if request.method == 'GET':
        code = request.GET.get('code', '')
        if code:
            product_set = Product.objects.filter(code=code)
            if product_set.exists():
                product_dic = []
                for p in product_set:
                    product = {
                        'id': p.id,
                        'code': p.code,
                        'name': str(p.name) + " \n" + str(p.measure())
                        # 'unit': dict(p.presentation_set.first()._meta.get_field('unit').choices)
                    }
                    product_dic.append(product)
                return JsonResponse({
                    'success': True,
                    'product': product_dic
                }, status=HTTPStatus.OK)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'El codigo no se encuentra registrado'
                }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'El codigo no se encuentra registrado'
            }, status=HTTPStatus.OK)


def get_purchase(request):
    if request.method == 'GET':
        number = request.GET.get('number', '')
        user_id = request.user.id
        user_obj = User.objects.get(id=user_id)
        subsidiary_obj = user_obj.subsidiary
        order_set = Order.objects.filter(type='C', subsidiary=subsidiary_obj, number=number)
        if order_set.exists():
            order_obj = order_set.last()
            if order_obj.status == 'P':
                detail = []
                order_detail_set = OrderDetail.objects.filter(order=order_obj, is_state=True)
                for d in order_detail_set.order_by('id'):
                    row = {
                        'id': d.id,
                        'code': d.product.code,
                        'product': d.product.id,
                        'description': d.product.name,
                        'units': d.get_unit_display(),
                        'unit': d.unit,
                        'quantity': d.quantity,
                        'quantity_niu': d.quantity_niu,
                        'price': d.price
                    }
                    detail.append(row)

                return JsonResponse({
                    'success': True,
                    'pk': order_obj.id,
                    'invoice': order_obj.invoice_number,
                    'discount': order_obj.total_discount,
                    'type': order_obj.doc,
                    'person': order_obj.person.id,
                    'person_document': order_obj.person.document,
                    'person_number': order_obj.person.number,
                    'person_names': order_obj.person.names,
                    'person_address': order_obj.person.address,
                    'person_email': order_obj.person.email,
                    'person_phone': order_obj.person.phone,
                    'date_document': order_obj.date_document,
                    'invoice_date': order_obj.invoice_date,
                    'coin': order_obj.coin,
                    'change': order_obj.change,
                    'igv': order_obj.is_igv,
                    'detail': detail,
                }, status=HTTPStatus.OK)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Compra ' + str(
                        order_obj.get_status_display()).lower() + ', solo las compras pendientes son modificables',
                }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'No existe la compra de ' + ' Nº ' + str(number),
            }, status=HTTPStatus.OK)


def valid_unit(request):
    if request.method == 'GET':
        quantity = request.GET.get('quantity', '')
        product = request.GET.get('product', '')
        unit = request.GET.get('unit', '')
        units = request.GET.get('units', '')
        presenting_set = None
        if quantity and product:
            product_obj = Product.objects.get(id=int(product))
            if unit == 'NIU':
                presenting_set = Presentation.objects.filter(product=product_obj, unit=unit, quantity_niu=int(quantity))
            elif unit == 'KGM':
                presenting_set = Presentation.objects.filter(product=product_obj, unit=unit)
            if presenting_set is None:
                return JsonResponse({
                    'success': False,
                    'message': 'Registre la presentacion con la unidad de medida en ' + str(units),
                }, status=HTTPStatus.OK)
            if presenting_set.exists():
                return JsonResponse({
                    'success': True,
                    'message': 'Ok'
                }, status=HTTPStatus.OK)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Registre la presentacion con la unidad de medida en ' + str(units) + ' y con la '
                                                                                                     'Unidad(uno) como principal',
                }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'Ingrese una cantidad',
            }, status=HTTPStatus.OK)


@csrf_exempt
def purchase_save(request):
    if request.method == 'POST':
        order_json = request.POST.get('order', '')
        data = json.loads(order_json)
        order_id = data['order-pk']
        document = data['document']
        creation = data['creation']
        invoice = data['invoice']
        date_document = data['date_document']
        invoice_date = data['invoice_date']
        coin = data['coin']
        change = data['change']
        igv = bool(int(data['igv']))
        state = 'P'
        date = datetime.strptime(creation, '%Y-%m-%d')
        types = data['type']
        person = data['person']
        person_obj = None
        if person:
            person_obj = Person.objects.get(id=int(person))
        total = data['total']
        total_discount = data['total_discount']
        user = request.user.id
        user_obj = User.objects.get(id=user)
        subsidiary_obj = user_obj.subsidiary
        pk = None
        if order_id != '0' and order_id != '':
            pk = int(order_id)
            obj = Order.objects.get(id=pk)
            correlative = obj.number
            purchase = obj.correlative
        else:
            correlative = number_order_purchase(subsidiary_obj, types)
            purchase = number_purchase(subsidiary_obj, types, date)
        order_obj, created = Order.objects.update_or_create(
            id=pk,
            defaults={
                "type": types,
                "doc": document,
                "number": correlative,
                "correlative": purchase,
                "status": state,
                "total": decimal.Decimal(total),
                "total_discount": decimal.Decimal(total_discount),
                "coin": coin,
                "change": decimal.Decimal(change),
                "create_at": date,
                "user": user_obj,
                "person": person_obj,
                "subsidiary": subsidiary_obj,
                "invoice_number": invoice,
                "invoice_date": invoice_date,
                "date_document": date_document,
                "is_igv": igv
            })
        if order_obj:
            for d in data['Detail']:
                detail = d['detail']
                product = d['product']
                product_obj = None
                if product != '0' or product != '':
                    product_obj = Product.objects.get(id=int(product))
                quantity = d['quantity']
                if quantity:
                    quantity = decimal.Decimal(quantity)
                else:
                    quantity = decimal.Decimal(0.00)
                q = d['q']
                if q:
                    q = int(q)
                else:
                    q = 0
                unit = d['unit']
                niu = int(minimum_quantity(product_obj, unit, quantity))
                price = d['price']
                dk = None
                if detail != '0' and detail != '':
                    dk = int(detail)
                OrderDetail.objects.update_or_create(
                    id=dk,
                    defaults={
                        "operation": 'E',
                        "order": order_obj,
                        "product": product_obj,
                        "quantity": decimal.Decimal(quantity),
                        "quantity_niu": niu,
                        "price": decimal.Decimal(price),
                        "unit": unit,
                        "is_state": True,
                        "is_invoice": False
                    })
            if order_obj:
                return JsonResponse({
                    'success': True,
                    'order': order_obj.id,
                    'code': str(order_obj.get_code()),
                    'message': 'Proceso realizado Codigo=' + str(order_obj.get_code())
                }, status=HTTPStatus.OK)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Ocurrio un problema en el proceso'
                }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'Ocurrio un problema en el proceso'
            }, status=HTTPStatus.OK)


def minimum_quantity(product=None, unit=None, quantity=0):
    if product and unit:
        if quantity:
            presenting_set = Presentation.objects.filter(product=product, unit=unit)
            if presenting_set.exists():
                if unit == 'NIU':
                    presenting_obj = presenting_set.filter(quantity_niu=1).first()
                    return round(quantity * decimal.Decimal(presenting_obj.quantity_niu))
                elif unit == 'KGM':
                    presenting_obj = presenting_set.first()
                    return round(decimal.Decimal(presenting_obj.quantity_niu) * quantity)
                elif unit == 'BG':
                    presenting_obj = presenting_set.first()
                    return round(decimal.Decimal(presenting_obj.quantity_niu) * quantity)
                elif unit == 'BX':
                    presenting_obj = presenting_set.first()
                    return round(decimal.Decimal(presenting_obj.quantity_niu) * quantity)
                else:
                    return 1
            else:
                return 0
        else:
            return 0
    else:
        return 0


def modal_kardex(request):
    if request.method == 'GET':
        product = request.GET.get('pk', '')
        if product:
            my_date = datetime.now()
            date_now = my_date.strftime("%Y-%m")
            product_obj = Product.objects.get(id=int(product))
            t = loader.get_template('sales/kardex_list.html')
            c = ({
                'product_obj': product_obj,
                'date_now': date_now
            })
            return JsonResponse({
                'success': True,
                'form': t.render(c, request),
            })


def modal_kardex_accounting(request):
    if request.method == 'GET':
        product = request.GET.get('pk', '')
        if product:
            my_date = datetime.now()
            date_now = my_date.strftime("%Y-%m-%d")
            product_obj = Product.objects.get(id=int(product))
            t = loader.get_template('sales/kardex.html')
            c = ({
                'product_obj': product_obj,
                'date_now': date_now
            })
            return JsonResponse({
                'success': True,
                'form': t.render(c, request),
            })


def get_kardex(request):
    if request.method == 'GET':
        init = request.GET.get('init', '')
        end = request.GET.get('end', '')
        product = request.GET.get('product', '')
        if init and end and product:
            product_obj = Product.objects.get(id=int(product))
            # m = month[5:]
            # detail_set = OrderDetail.objects.filter(product=product_obj, is_state=True,
            #                                         order__create_at__month=int(m)).order_by('id')
            order_detail_obj = OrderDetail.objects.filter(product=product_obj, order__type__in=['V', 'C']).first()
            i_d = 0
            if order_detail_obj:
                i_d = order_detail_obj.id
            detail_set = OrderDetail.objects.filter(product=product_obj,
                                                    order__create_at__range=(init, end),
                                                    order__type__in=['V', 'C']).order_by('id')
            tpl = loader.get_template('sales/kardex_grid_list.html')
            context = ({
                'detail_set': detail_set,
                'product_obj': product_obj,
                'id': i_d
            })
            return JsonResponse({
                'grid': tpl.render(context, request),
            }, status=HTTPStatus.OK, content_type="application/json")


def get_search_product(request):
    if request.method == 'GET':
        product = request.GET.get('product', '')
        if product:
            product_set = Product.objects.filter(
                name__contains=product.upper()).select_related('brand', 'family').prefetch_related(
                Prefetch(
                    'presentation_set', queryset=Presentation.objects.select_related('product')
                )
            ).order_by('id')
            tpl = loader.get_template('sales/product_grid_list.html')
            context = ({
                'product_set': product_set
            })
            return JsonResponse({
                'grid': tpl.render(context, request),
            }, status=HTTPStatus.OK, content_type="application/json")


def get_search_product_other(request):
    if request.method == 'GET':
        name = request.GET.get('name', '')
        other = request.GET.get('other', '')
        if name:
            if other == '1':
                product_set = Product.objects.filter(brand__name__contains=name.upper()).order_by('id')
            elif other == '2':
                product_set = Product.objects.filter(family__name__contains=name.upper()).order_by('id')
            else:
                product_set = []
            tpl = loader.get_template('sales/product_grid_list.html')
            context = ({
                'product_set': product_set
            })
            return JsonResponse({
                'grid': tpl.render(context, request),
            }, status=HTTPStatus.OK, content_type="application/json")


def get_search_product_code(request):
    if request.method == 'GET':
        code = request.GET.get('code', '')
        if code:
            product_set = Product.objects.filter(code=code).order_by('id')
            tpl = loader.get_template('sales/product_grid_list.html')
            context = ({
                'product_set': product_set
            })
            return JsonResponse({
                'success': True,
                'grid': tpl.render(context, request),
            }, status=HTTPStatus.OK, content_type="application/json")
        else:
            return JsonResponse({
                'success': False,
                'message': 'Problemas con el codigo'
            }, status=HTTPStatus.OK)


def get_product_all(request):
    if request.method == 'GET':
        product_set = Product.objects.all().select_related('family', 'brand').order_by('id')
        tpl = loader.get_template('sales/product_grid_list.html')
        context = ({
            'product_set': product_set
        })
        return JsonResponse({
            'success': True,
            'grid': tpl.render(context, request),
        }, status=HTTPStatus.OK, content_type="application/json")


def get_change(request):
    if request.method == 'GET':
        date = request.GET.get('date', '')
        # my_date = datetime.now()
        # date_now = my_date.strftime("%Y-%m-%d")
        c = api_net.get_exchange_rate(str(date))
        if c:
            if c.get('success') is True:
                sales = c.get('venta')
                money = c.get('moneda')
                return JsonResponse({
                    'success': True,
                    'sales': sales,
                    'money': money
                }, status=HTTPStatus.OK)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Problemas con el tipo cambio'
                }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'No se obtuvieron resultados'
            }, status=HTTPStatus.OK)


def graphic_user(request):
    if request.method == 'GET':
        date_now = datetime.now().date()
        user_set = User.objects.filter(type__in=['V', 'C', 'A'], is_active=True)
        label = []
        sales = []
        month = []
        for u in user_set:
            label.append(u.first_name.upper())
            order_set = Order.objects.filter(create_at=date_now, type='V')
            month_set = Order.objects.filter(create_at__month=date_now.month, create_at__year=date_now.year, type='V')
            total_m = month_set.filter(payments__isnull=False, user=u).distinct('id').values('id', 'total')
            total_v = order_set.filter(payments__isnull=False, user=u).distinct('id').values('id', 'total')
            t_v = decimal.Decimal(0.00)
            for i in total_v:
                t_v += decimal.Decimal(i['total'])
            t_m = decimal.Decimal(0.00)
            for m in total_m:
                t_m += decimal.Decimal(m['total'])
            sales.append(round(float(t_v), 2))
            month.append(round(float(t_m), 2))

        return JsonResponse({
            'sales': sales,
            'month': month,
            'label': label
        }, status=HTTPStatus.OK)


def graphic_sales(request):
    if request.method == 'GET':
        date_now = datetime.now().date()
        order_set = Order.objects.filter(create_at=date_now, type='V')
        total_f = order_set.filter(doc='1', payments__isnull=False).distinct('id').values('id', 'total')
        t_f = decimal.Decimal(0.00)
        for i in total_f:
            t_f += decimal.Decimal(i['total'])
        total_b = order_set.filter(doc='2', payments__isnull=False).distinct('id').values('id', 'total')
        t_b = decimal.Decimal(0.00)
        for i in total_b:
            t_b += decimal.Decimal(i['total'])
        total_t = order_set.filter(doc='0', payments__isnull=False).distinct('id').values('id', 'total')
        t_t = decimal.Decimal(0.00)
        for i in total_t:
            t_t += decimal.Decimal(i['total'])
        return JsonResponse({
            'total': [t_t, t_f, t_b],
            'label': ['TICKETS', 'FACTURAS', 'BOLETAS']
        }, status=HTTPStatus.OK)


def graphic_month(request):
    if request.method == 'GET':
        date_now = datetime.now().date()
        data = []
        i = 1
        while i < 13:
            order_set = Order.objects.filter(create_at__month=i, create_at__year=date_now.year, type='V')
            total_sales = order_set.filter(payments__isnull=False).distinct('id').values('id', 'total')
            sales = decimal.Decimal(0.00)
            for f in total_sales:
                sales += decimal.Decimal(f['total'])
            i += 1
            data.append(sales)

        return JsonResponse({
            'd': data,
            'label': ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Setiembre', 'Octubre',
                      'Noviembre', 'Diciembre']
        }, status=HTTPStatus.OK)


def get_sales_detail(request):
    if request.method == 'GET':
        pk = request.GET.get('pk', '')
        order_obj = Order.objects.get(pk=int(pk))
        order_detail_set = OrderDetail.objects.filter(order=order_obj, is_state=True)
        t = loader.get_template('accounting/orders_detail_grid.html')
        c = ({
            'order_detail_set': order_detail_set,
        })
        return JsonResponse({
            'order': order_obj.number,
            'grid': t.render(c, request),
        }, status=HTTPStatus.OK)


@csrf_exempt
def export_product(request):
    if request.method == 'POST':
        product_json = request.POST.get('product', '')
        detail = json.loads(product_json)
        document = detail['name']
        if document:
            wb = Workbook()
            bandera = True
            cont = 1
            row = 4
            if bandera:
                ws = wb.active
                ws.title = str(document)  # hoja
                bandera = False
            else:
                ws = wb.create_sheet(str(document) + str(cont))
            # Crear el título en la hoja
            my_date = datetime.now()
            date_now = my_date.strftime("%d-%m-%Y")
            hour_now = my_date.strftime("%H:%M:%S")
            ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B1'].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type="solid")
            ws['B1'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
            ws['B1'] = 'PRODUCTOS CON STOCK AL MINIMO ' + ' FECHA: ' + str(date_now) + ' HORA: ' + str(hour_now)

            # Cambiar caracteristicas de las celdas
            ws.merge_cells('B1:N1')

            ws.row_dimensions[1].height = 25

            ws.column_dimensions['B'].width = 9
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 9
            ws.column_dimensions['F'].width = 9
            ws.column_dimensions['G'].width = 9
            ws.column_dimensions['H'].width = 9
            ws.column_dimensions['I'].width = 9
            ws.column_dimensions['J'].width = 9
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 9
            ws.column_dimensions['M'].width = 9
            ws.column_dimensions['N'].width = 25

            # Crear la cabecera
            ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['B3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['B3'] = 'CODIGO'

            ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['C3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['C3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['C3'] = 'NOMBRE PRODUCTO'

            ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['D3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['D3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['D3'] = 'DESCRIPCION PRODUCTO'

            ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['E3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['E3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['E3'] = 'MARCA'

            ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['F3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['F3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['F3'] = 'FAMILIA'

            ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['G3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['G3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['G3'] = 'TIPO'

            ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['H3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['H3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['H3'] = 'ANCHO'

            ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['I3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['I3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['I3'] = 'LARGO'

            ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['J3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['J3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['J3'] = 'ALTO'

            ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['K3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['K3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['K3'] = 'ALMACEN'

            ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['L3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['L3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['L3'] = 'STOCK'

            ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['M3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['M3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['M3'] = 'MINIMO'

            ws['N3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['N3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['N3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            ws['N3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
            ws['N3'] = 'PRODUCTO RELACIONADO'

            for p in detail['Detail']:
                # Pintamos los datos en el reporte
                color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                stock_cell = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type="solid")
                text_color = Font(name='Calibri', size=9)
                border_cell = Border(left=Side(border_style="thin"),
                                     right=Side(border_style="thin"),
                                     top=Side(border_style="thin"),
                                     bottom=Side(border_style="thin"))
                align_cell = Alignment(horizontal="center")
                ws.cell(row=row, column=2).alignment = align_cell
                ws.cell(row=row, column=2).border = border_cell
                ws.cell(row=row, column=2).fill = color_cell
                ws.cell(row=row, column=2).font = text_color
                ws.cell(row=row, column=2).value = p['code']

                ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
                ws.cell(row=row, column=3).border = border_cell
                ws.cell(row=row, column=3).fill = color_cell
                ws.cell(row=row, column=3).font = text_color
                ws.cell(row=row, column=3).value = p['product']

                ws.cell(row=row, column=4).alignment = Alignment(horizontal="left")
                ws.cell(row=row, column=4).border = border_cell
                ws.cell(row=row, column=4).fill = color_cell
                ws.cell(row=row, column=4).font = text_color
                ws.cell(row=row, column=4).value = ""

                ws.cell(row=row, column=5).alignment = align_cell
                ws.cell(row=row, column=5).border = border_cell
                ws.cell(row=row, column=5).fill = color_cell
                ws.cell(row=row, column=5).font = text_color
                ws.cell(row=row, column=5).value = p['brand']

                ws.cell(row=row, column=6).alignment = align_cell
                ws.cell(row=row, column=6).border = border_cell
                ws.cell(row=row, column=6).fill = color_cell
                ws.cell(row=row, column=6).font = text_color
                ws.cell(row=row, column=6).value = p['family']

                ws.cell(row=row, column=7).alignment = align_cell
                ws.cell(row=row, column=7).border = border_cell
                ws.cell(row=row, column=7).fill = color_cell
                ws.cell(row=row, column=7).font = text_color
                ws.cell(row=row, column=7).value = p['type']

                ws.cell(row=row, column=8).alignment = align_cell
                ws.cell(row=row, column=8).border = border_cell
                ws.cell(row=row, column=8).fill = color_cell
                ws.cell(row=row, column=8).font = text_color
                ws.cell(row=row, column=8).value = p['width']

                ws.cell(row=row, column=9).alignment = align_cell
                ws.cell(row=row, column=9).border = border_cell
                ws.cell(row=row, column=9).fill = color_cell
                ws.cell(row=row, column=9).font = text_color
                ws.cell(row=row, column=9).value = p['length']

                ws.cell(row=row, column=10).alignment = align_cell
                ws.cell(row=row, column=10).border = border_cell
                ws.cell(row=row, column=10).fill = color_cell
                ws.cell(row=row, column=10).font = text_color
                ws.cell(row=row, column=10).value = p['height']

                ws.cell(row=row, column=11).alignment = align_cell
                ws.cell(row=row, column=11).border = border_cell
                ws.cell(row=row, column=11).fill = color_cell
                ws.cell(row=row, column=11).font = text_color
                ws.cell(row=row, column=11).value = p['store']

                ws.cell(row=row, column=12).alignment = align_cell
                ws.cell(row=row, column=12).border = border_cell
                ws.cell(row=row, column=12).fill = stock_cell
                ws.cell(row=row, column=12).font = text_color
                ws.cell(row=row, column=12).value = p['stock']

                ws.cell(row=row, column=13).alignment = align_cell
                ws.cell(row=row, column=13).border = border_cell
                ws.cell(row=row, column=13).fill = color_cell
                ws.cell(row=row, column=13).font = text_color
                ws.cell(row=row, column=13).value = p['minimum']

                ws.cell(row=row, column=14).alignment = align_cell
                ws.cell(row=row, column=14).border = border_cell
                ws.cell(row=row, column=14).fill = color_cell
                ws.cell(row=row, column=14).font = text_color
                ws.cell(row=row, column=14).value = p['relation']

                cont += 1
                row += 1

            # Establecer el nombre de mi archivo
            nombre_archivo = "ProductFilter.xlsx"
            # Definir el tipo de respuesta que se va a dar
            response = HttpResponse(content_type="application/ms-excel")
            contenido = "attachment; filename = {0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response


def purchase_list(request):
    if request.method == 'GET':
        my_date = datetime.now()
        return render(request, 'sales/purchase_list.html', {
            'date_now': my_date.date()
        })


def get_purchase_by_date(request):
    if request.method == 'GET':
        init = request.GET.get('init', '')
        end = request.GET.get('end', '')
        if init and end:
            order_set = Order.objects.filter(create_at__range=(init, end), type='C').order_by('id')
            tpl = loader.get_template('sales/purchase_grid_list.html')
            context = ({'order_set': order_set})
            return JsonResponse({
                'grid': tpl.render(context, request),
            }, status=HTTPStatus.OK, content_type="application/json")


def pass_purchase(request):
    if request.method == 'GET':
        order = request.GET.get('order', '')
        if order:
            order_obj = Order.objects.get(id=int(order))
            for d in order_obj.orderdetail_set.filter(is_state=True).order_by('id'):
                if d:
                    input_stock(d)
                    # if order_obj.doc in ['1', '2']:
                    #     if not Kardex.objects.filter(order_detail=d).exists():
                    #         kardex_set = Kardex.objects.filter(product=d.product)
                    #         if kardex_set.exists():
                    #             type_document = '00'
                    #             if order_obj.doc == '1':
                    #                 type_document = str('0' + order_obj.doc)
                    #             elif order_obj.doc == '2':
                    #                 type_document = '03'
                    #             kardex_input(product=d.product, quantity=d.quantity, total_cost=d.amount(),
                    #                          order_detail_obj=d, type_document=type_document,
                    #                          type_operation='02')
                    #         else:
                    #             kardex_initial(product=d.product, stock=d.quantity, price_unit=d.price,
                    #                            order_detail_obj=d)
            if order_obj:
                order_obj.status = 'R'
                order_obj.save()
            return JsonResponse({'success': True, 'message': 'Orden aprobada'}, status=HTTPStatus.OK)
        else:
            return JsonResponse({'success': False, 'message': 'Orden sin identificar'}, status=HTTPStatus.OK)


def cancel_purchase(request):
    if request.method == 'GET':
        order = request.GET.get('order', '')
        if order:
            order_obj = Order.objects.get(id=int(order))
            for d in order_obj.orderdetail_set.filter(is_state=True).order_by('id'):
                if d:
                    output_stock(d)
                    # d.operation = 'S'   VALORES QUE DEBERIAN ESTAR REVISAR
                    # d.save()

                    # kardex_set = Kardex.objects.filter(product=d.product)
                    # if kardex_set.exists():
                    #     kardex_ouput(product=d.product, quantity=d.quantity, order_detail_obj=d,
                    #                  type_document=str('0' + order_obj.doc), type_operation='05')

            if order_obj:
                order_obj.status = 'A'
                order_obj.save()
            return JsonResponse({'success': True, 'message': 'Orden Anulada'}, status=HTTPStatus.OK)
        else:
            return JsonResponse({'success': False, 'message': 'Orden sin identificar'}, status=HTTPStatus.OK)


@csrf_exempt
def quotation_save(request):
    if request.method == 'POST':
        order_json = request.POST.get('order', '')
        data = json.loads(order_json)
        order_id = data['order-pk']
        document = data['document']
        state = 'R'
        date = datetime.now().date()
        current_time = datetime.now()
        types = data['type']
        person = data['person']
        license_plate = data['license_plate']
        person_obj = None
        if person:
            person_obj = Person.objects.get(id=int(person))
        total = data['total']
        total_discount = data['total_discount']
        user = request.user.id
        user_obj = User.objects.get(id=user)
        subsidiary_obj = user_obj.subsidiary
        pk = None
        if order_id != '0' and order_id != '':
            pk = int(order_id)
            obj = Order.objects.get(id=pk)
            correlative = obj.number
        else:
            last_order_set = Order.objects.filter(user=user_obj)
            if last_order_set.exists():
                last_order_obj = last_order_set.last()
                print("current_time", current_time.timestamp())
                print("last_order_obj.update_at", last_order_obj.update_at.timestamp())
                time_difference = current_time.timestamp() - last_order_obj.update_at.timestamp()
                print(time_difference)
                if time_difference >= 15:
                    correlative = number_order(subsidiary_obj, types)
                else:
                    if order_id != '0' and order_id != '':
                        pk = int(order_id)
                        obj = Order.objects.get(id=pk)
                        correlative = obj.number
                    else:
                        data = {'error': "Orden Registrada"}
                        response = JsonResponse(data)
                        response.status_code = HTTPStatus.INTERNAL_SERVER_ERROR
                        return response
            else:
                correlative = number_order(subsidiary_obj, types)
        order_obj, created = Order.objects.update_or_create(
            id=pk,
            defaults={
                "type": types,
                "doc": document,
                "number": correlative,
                "status": state,
                "total": decimal.Decimal(total),
                "total_discount": decimal.Decimal(total_discount),
                "coin": '1',
                "change": 1,
                "create_at": date,
                "user": user_obj,
                "person": person_obj,
                "subsidiary": subsidiary_obj,
                "license_plate": license_plate
            })
        if order_obj:
            for d in data['Detail']:
                detail = d['detail']
                product = d['product']
                product_obj = None
                if product != '0' or product != '':
                    product_obj = Product.objects.get(id=int(product))
                quantity = d['quantity']
                q = d['q']
                if q:
                    q = int(q)
                else:
                    q = 0
                unit = d['unit']
                niu = d['niu']
                units = decimal.Decimal(quantity)
                if unit == 'KGM':
                    units = round(decimal.Decimal(quantity) * decimal.Decimal(niu))
                price = decimal.Decimal(d['price'])
                is_sate = False
                operation = None
                invoice = False
                if types == 'T':
                    operation = 'S'
                    is_sate = False
                    invoice = False
                elif types == 'V':
                    operation = 'S'
                    is_sate = True
                    invoice = True
                elif types == 'C':
                    operation = 'E'
                    is_sate = True
                    invoice = False
                dk = None
                if detail != '0' and detail != '':
                    dk = int(detail)
                order_detail_obj, created_detail = OrderDetail.objects.update_or_create(
                    id=dk,
                    defaults={
                        "operation": operation,
                        "order": order_obj,
                        "product": product_obj,
                        "quantity": decimal.Decimal(quantity),
                        "quantity_niu": int(units),
                        "price": decimal.Decimal(round(price, 4)),
                        "unit": unit,
                        "is_state": is_sate,
                        "is_invoice": invoice
                    })
            if order_obj:
                if user_obj.is_authorization:
                    user_obj.is_authorization = False
                    user_obj.save()
                return JsonResponse({
                    'success': True,
                    'order': order_obj.id,
                    'number': order_obj.number,
                    'hatch': order_obj.user.user_work,  # ventanilla
                    'userID': order_obj.user.id,  # ventanilla
                    'message': 'Cotización realizada'
                }, status=HTTPStatus.OK)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Ocurrio un problema en el proceso'
                }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'Ocurrio un problema en el proceso'
            }, status=HTTPStatus.OK)
    return JsonResponse({'message': 'Error de peticion.'}, status=HTTPStatus.BAD_REQUEST)


def get_kardex_table(request):
    if request.method == 'GET':
        init = request.GET.get('init', '')
        end = request.GET.get('end', '')
        product = request.GET.get('product', '')
        if init and end and product:
            product_obj = Product.objects.get(id=int(product))
            # m = month[5:]
            # detail_set = OrderDetail.objects.filter(product=product_obj, is_state=True,
            #                                         order__create_at__month=int(m)).order_by('id')
            order_detail_obj = OrderDetail.objects.filter(product=product_obj, order__type__in=['V', 'C']).first()
            i_d = 0
            if order_detail_obj:
                i_d = order_detail_obj.id
            detail_set = OrderDetail.objects.filter(product=product_obj,
                                                    create_at__date__range=(init, end),
                                                    order__type__in=['V', 'C']).order_by('create_at')
            tpl = loader.get_template('sales/kardex_table.html')
            context = ({
                'detail_set': detail_set,
                'product_obj': product_obj,
                'id': i_d
            })
            return JsonResponse({
                'grid': tpl.render(context, request),
            }, status=HTTPStatus.OK, content_type="application/json")


def get_presentation(request):
    if request.method == 'GET':
        p = request.GET.get('pk', '')
        if p:
            presentation_obj = Presentation.objects.get(id=int(p))
            return JsonResponse({
                'success': True,
                'pk': presentation_obj.pk,
                'quantity': presentation_obj.quantity,
                'quantity_niu': presentation_obj.quantity_niu,
                'price': presentation_obj.price,
                'unit': presentation_obj.unit,
                'is_corporate': presentation_obj.is_corporate
            }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'No existe la presentacion',
            }, status=HTTPStatus.OK)


def update_presentation(request):
    if request.method == 'GET':
        p = request.GET.get('pk', '')
        quantity = request.GET.get('quantity', '')
        quantity_niu = request.GET.get('quantity_niu', '')
        unit = request.GET.get('unit', '')
        price = request.GET.get('price', '')
        is_corporate = request.GET.get('is_corporate', '')
        _is_corporate = False
        if is_corporate != '0':
            _is_corporate = True
        if p:
            presentation_obj = Presentation.objects.get(id=int(p))
            presentation_obj.quantity = decimal.Decimal(quantity)
            presentation_obj.quantity_niu = decimal.Decimal(quantity_niu)
            presentation_obj.price = decimal.Decimal(price)
            presentation_obj.unit = unit
            presentation_obj.is_corporate = _is_corporate
            presentation_obj.save()
            return JsonResponse({
                'success': True,
                'pk': presentation_obj.pk,
                'quantity': presentation_obj.quantity,
                'quantity_niu': presentation_obj.quantity_niu,
                'price': presentation_obj.price,
                'unit': presentation_obj.get_unit_display(),
                'is_corporate': presentation_obj.is_corporate
            }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'No existe la presentacion',
            }, status=HTTPStatus.OK)


class SalesListCorporate(ListView):
    model = Product
    template_name = 'sales/order_corporate.html'

    def get_context_data(self, **kwargs):
        context = {
            'document_set': Person._meta.get_field('document').choices,
            'order_type_set': Order._meta.get_field('type').choices,
            'doc_set': Order._meta.get_field('doc').choices,
        }
        return context


def cancel_order(request):
    if request.method == 'GET':
        order = request.GET.get('order', '')
        if order:
            order_obj = Order.objects.get(id=int(order))
            if order_obj.doc == '1' or order_obj.doc == '2':
                order_detail_set = order_obj.orderdetail_set.filter(is_state=True)
                for d in order_detail_set:
                    new_order_detail = {
                        "operation": 'E',
                        "order": order_obj,
                        "product": d.product,
                        "quantity": decimal.Decimal(d.quantity),
                        "quantity_niu": decimal.Decimal(d.quantity_niu),
                        "price": decimal.Decimal(round(d.price, 6)),
                        "unit": d.unit,
                        "is_state": False,
                        "is_invoice": False
                    }
                    new_order_detail_obj = OrderDetail.objects.create(**new_order_detail)
                    new_order_detail_obj.save()
                    input_stock(order_detail_obj=new_order_detail_obj)
                    d.is_state = False
                    d.save()
                if order_obj.bill_type and order_obj.bill_number:
                    order_obj.condition = 'PA'
                    order_obj.save()
                else:
                    order_obj.condition = 'A'
                    order_obj.status = 'A'
                    order_obj.save()
                return JsonResponse({
                    'success': True,
                    'last_order': order_obj.id,
                    'number': order_obj.number,
                    'message': 'Comprobante Electronico Anulado correctamente',
                }, status=HTTPStatus.OK)
            elif order_obj.doc == '0':
                order_detail_set = order_obj.orderdetail_set.filter(is_state=True)
                for d in order_detail_set:
                    d.is_state = False
                    d.save()
                    new_order_detail = {
                        "operation": 'E',
                        "order": order_obj,
                        "product": d.product,
                        "quantity": decimal.Decimal(d.quantity),
                        "quantity_niu": int(d.quantity_niu),
                        "price": decimal.Decimal(round(d.price, 6)),
                        "unit": d.unit,
                        "is_state": False,
                        "is_invoice": False
                    }
                    new_order_detail_obj = OrderDetail.objects.create(**new_order_detail)
                    new_order_detail_obj.save()
                    input_stock(order_detail_obj=new_order_detail_obj)
                    order_obj.status = 'A'
                    order_obj.condition = 'A'
                    order_obj.bill_number = None
                    order_obj.bill_serial = None
                    order_obj.bill_qr = None
                    order_obj.bill_date = None
                    order_obj.bill_hash = None
                    order_obj.bill_type = None
                    order_obj.bill_status = None
                    order_obj.bill_enlace_pdf = None
                    order_obj.save()
                return JsonResponse({
                    'success': True,
                    'last_order': order_obj.id,
                    'number': order_obj.number,
                    'message': 'Orden anulado correctamente',
                }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'Orden sin identificar',
            }, status=HTTPStatus.OK)
    return JsonResponse({'message': 'Error de peticion.'}, status=HTTPStatus.BAD_REQUEST)


def get_last_order(request):
    if request.method == 'GET':
        number = request.GET.get('number', '')
        order_type = request.GET.get('type', '')
        type_name = request.GET.get('type_name', '')
        user_id = request.user.id
        user_obj = User.objects.get(id=user_id)
        subsidiary_obj = user_obj.subsidiary
        order_set = Order.objects.filter(type=order_type, subsidiary=subsidiary_obj, number=number)
        if order_set.exists():
            order_obj = order_set.last()
            detail = []
            order_detail_set = OrderDetail.objects.filter(order=order_obj, is_state=False, operation='S')
            person = ''
            person_number = ''
            person_names = ''
            person_address = ''
            relation = None
            for d in order_detail_set.order_by('id'):
                y = 0
                if d.product.relation != '' and d.product.relation != '-':
                    relation = d.product.relation
                if relation == d.product.code:
                    y = 1
                    relation = None
                row = {
                    'id': d.id,
                    'code': d.product.code,
                    'product': d.product.id,
                    'description': d.product.name,
                    'measure': d.product.measure(),
                    'unit': d.unit,
                    'quantity': d.quantity,
                    'quantity_niu': d.quantity_niu,
                    'price': d.price,
                    'm': get_mayor(d.product),
                    'y': y,
                    'order_id': d.order.id,
                    'stock': d.product.stock
                }
                detail.append(row)
                if order_obj.person:
                    person = order_obj.person.id
                    person_number = order_obj.person.number,
                    person_names = order_obj.person.names,
                    person_address = order_obj.person.address,
            return JsonResponse({
                'success': True,
                'pk': order_obj.id,
                'last_order': order_obj.id,
                'type': order_obj.type,
                'discount': order_obj.total_discount,
                'license_plate': order_obj.license_plate,
                'doc': order_obj.doc,
                'person': person,
                'person_number': person_number,
                'person_names': person_names,
                'person_address': person_address,
                'detail': detail,
            }, status=HTTPStatus.OK)
        else:
            return JsonResponse({
                'success': False,
                'message': 'No existe la Orden de ' + str(type_name).capitalize() + ' Nº ' + str(number),
            }, status=HTTPStatus.OK)


def kardex_initial(
        product,
        stock,
        price_unit,
        order_detail_obj=None,
        create_at=None,
        quantity_in_unit=0
):
    if create_at is None:
        create_at = datetime.now()

    remaining_price_total = decimal.Decimal(stock) * decimal.Decimal(price_unit)

    if quantity_in_unit != 0:
        stock = quantity_in_unit

    Kardex.objects.create(
        operation='C',
        quantity=0,
        price_unit=0,
        price_total=0,
        remaining_quantity=decimal.Decimal(stock),
        remaining_price=decimal.Decimal(price_unit),
        remaining_price_total=remaining_price_total,
        order_detail=order_detail_obj,
        type_document='00',
        type_operation='16',
        product=product,
        create_at=create_at
    )
    product.invoice_stock = stock
    product.save()


def kardex_ouput(
        product,
        quantity,
        order_detail_obj=None,
        type_document='00',
        type_operation='99',
        create_at=None,
        quantity_in_unit=0
):
    old_stock = product.invoice_stock
    last_kardex = Kardex.objects.filter(product=product).last()
    last_remaining_quantity = last_kardex.remaining_quantity
    old_price_unit = last_kardex.remaining_price

    if quantity_in_unit != 0:
        new_remaining_quantity = last_remaining_quantity - quantity_in_unit
        quantity = quantity_in_unit
    else:
        new_remaining_quantity = last_remaining_quantity - quantity

    total_cost = decimal.Decimal(quantity) * old_price_unit
    new_remaining_price = last_kardex.remaining_price
    new_remaining_price_total = new_remaining_quantity * new_remaining_price
    new_stock = old_stock - decimal.Decimal(quantity)

    if create_at is None:
        create_at = datetime.now()

    Kardex.objects.create(
        operation='S',
        quantity=quantity,
        price_unit=old_price_unit,
        price_total=total_cost,
        remaining_quantity=new_remaining_quantity,
        remaining_price=new_remaining_price,
        remaining_price_total=new_remaining_price_total,
        order_detail=order_detail_obj,
        type_document=type_document,
        type_operation=type_operation,
        product=product,
        create_at=create_at
    )

    product.invoice_stock = new_stock
    product.save()


def kardex_input(
        product,
        quantity,
        total_cost,
        order_detail_obj=None,
        type_document='00',
        type_operation='99',
        create_at=None,
        quantity_in_unit=0
):
    old_stock = product.invoice_stock
    last_kardex = Kardex.objects.filter(product=product).last()
    last_remaining_quantity = last_kardex.remaining_quantity
    price_unit = decimal.Decimal(total_cost) / decimal.Decimal(quantity)

    last_remaining_price_total = last_kardex.remaining_price_total

    if quantity_in_unit != 0:
        new_remaining_quantity = last_remaining_quantity + quantity_in_unit
        quantity = quantity_in_unit
    else:
        new_remaining_quantity = last_remaining_quantity + quantity

    if new_remaining_quantity > 0:
        new_remaining_price = (decimal.Decimal(last_remaining_price_total) + total_cost) / new_remaining_quantity
    else:
        new_remaining_price = 0

    new_stock = old_stock + decimal.Decimal(quantity)
    new_remaining_price_total = new_remaining_quantity * new_remaining_price

    if create_at is None:
        create_at = datetime.now()

    Kardex.objects.create(
        operation='E',
        quantity=quantity,
        price_unit=price_unit,
        price_total=total_cost,
        remaining_quantity=new_remaining_quantity,
        remaining_price=new_remaining_price,
        remaining_price_total=new_remaining_price_total,
        order_detail=order_detail_obj,
        type_document=type_document,
        type_operation=type_operation,
        product=product,
        create_at=create_at
    )

    product.invoice_stock = new_stock
    product.save()


def get_kardex_by_product(request):
    user_id = request.user.id
    user_obj = User.objects.get(id=user_id)
    my_date = datetime.now()
    formatdate = my_date.strftime("%Y-%m")
    if request.method == 'GET':
        return render(request, 'sales/new_kardex_list.html', {
            'date_now': formatdate,
        })

    elif request.method == 'POST':
        product_code = request.POST.get('product-code')
        date_product = request.POST.get('date-product')
        product_obj = Product.objects.get(code=product_code)
        year, month = map(int, date_product.split('-'))
        start_date = datetime(year, month, 1)

        last_day_of_month = calendar.monthrange(year, month)[1]
        end_date = datetime(year, month, last_day_of_month)

        start_date = timezone.make_aware(start_date)
        end_date = timezone.make_aware(end_date)

        kardex_dict = []

        kardex_set = Kardex.objects.filter(product=product_obj,
                                           create_at__range=(start_date, end_date)
                                           ).select_related('order_detail', 'product').order_by('id')

        # kardex_set = Kardex.objects.filter(product=product_obj)

        sum_quantities_entries = 0
        sum_total_cost_entries = 0
        sum_quantities_exits = 0
        sum_total_cost_exits = 0

        sum_remaining_price = 0
        last_remaining_quantity = 0
        last_remaining_price_total = 0

        if kardex_set.exists():

            last_month = month - 1 if month > 1 else 12

            last_month_data_query = Kardex.objects.filter(
                product=product_obj,
                create_at__month=last_month
            ).order_by('-create_at', '-id').values('id', 'remaining_quantity', 'remaining_price_total').first()

            if last_month_data_query:
                last_month_id = last_month_data_query['id'] if last_month_data_query else None
                last_month_remaining_quantity = last_month_data_query[
                    'remaining_quantity'] if last_month_data_query else None
                last_month_remaining_price_total = last_month_data_query[
                    'remaining_price_total'] if last_month_data_query else None

            else:
                initial_operation = next((k for k in kardex_set if k.operation == 'C'), None)
                if initial_operation:
                    last_month_id = initial_operation.id
                    last_month_remaining_quantity = initial_operation.remaining_quantity
                    last_month_remaining_price_total = initial_operation.remaining_price_total
                else:
                    previous_movements = Kardex.objects.filter(
                        product=product_obj,
                        create_at__lt=start_date
                    ).order_by('-create_at')  

                    last_previous_movement = previous_movements.first()
                    if last_previous_movement:
                        last_month_id = last_previous_movement.id
                        last_month_remaining_quantity = last_previous_movement.remaining_quantity
                        last_month_remaining_price_total = last_previous_movement.remaining_price_total
                    else:
                        last_month_id = None
                        last_month_remaining_quantity = 0
                        last_month_remaining_price_total = 0

            # last_average_cost = last_month_remaining_price_total / last_month_remaining_quantity
            try:
                if last_month_remaining_quantity == 0:
                    last_average_cost = decimal.Decimal('0')
                else:
                    last_average_cost = last_month_remaining_price_total / last_month_remaining_quantity
            except DivisionUndefined:
                last_average_cost = decimal.Decimal('0')
            remaining_quantity = decimal.Decimal(last_month_remaining_quantity)
            remaining_price_total = decimal.Decimal(last_month_remaining_price_total)
            first_date = datetime.strptime(date_product + '-01', "%Y-%m-%d")

            item = {
                'id': last_month_id,
                'operation': 'C',
                'date': first_date.strftime("%d/%m/%Y"),
                'type_document': '00',
                'serial': 'INICIAL',
                'number': '0',
                'detail': 'SALDO INICIAL',
                'type_operation': '16',
                'quantity': str(round(last_month_remaining_quantity, 2)),
                'unit_cost': '',
                'total_cost': str(round(last_month_remaining_price_total, 2)),
                'remaining_quantity': str(round(last_month_remaining_quantity, 2)),
                'remaining_price': str(round(last_average_cost, 4)),
                'remaining_price_total': str(round(last_month_remaining_price_total, 2))
            }
            kardex_dict.append(item)

            for k in kardex_set:
                if k.operation == 'C':
                    continue
                serial = ''
                number = ''
                detail = ''
                if k.type_operation == '02':
                    nro_document = k.order_detail.order.invoice_number
                    serial, number = nro_document.split('-')
                    detail = 'COMPRA'
                elif k.type_operation == '01':
                    serial = k.order_detail.order.bill_serial
                    number = k.order_detail.order.bill_number
                    detail = 'VENTAS DEL DÍA'
                elif k.type_operation == '06':
                    serial = k.order_detail.order.note_serial
                    number = k.order_detail.order.note_number
                    detail = 'NOTA DE CREDITO'
                elif k.type_operation == '05':
                    nro_document = k.order_detail.order.invoice_number
                    serial, number = nro_document.split('-')
                    detail = 'NC (COMPRA)'

                if k.operation == 'E':
                    remaining_quantity += k.quantity
                    remaining_price_total += k.price_total
                    sum_quantities_entries += k.quantity
                    sum_total_cost_entries += k.price_total
                elif k.operation == 'S':
                    remaining_quantity -= k.quantity
                    remaining_price_total -= k.price_total
                    sum_quantities_exits += k.quantity
                    sum_total_cost_exits += k.price_total

                remaining_price = remaining_price_total / remaining_quantity if remaining_quantity != 0 else 0

                sum_remaining_price += remaining_price

                item = {
                    'id': k.id,
                    'product': k.product.id,
                    'operation': k.operation,
                    'order': k.order_detail.order,
                    'period': k.create_at.strftime("%Y-%m"),
                    'date': k.create_at.strftime("%d/%m/%Y"),
                    'type_document': k.type_document,
                    'serial': serial,
                    'number': number,
                    'detail': detail,
                    'type_operation': k.type_operation,
                    'quantity': k.quantity,
                    'unit_cost': round(k.price_unit, 2),
                    'total_cost': str(round(k.price_total, 2)),
                    'remaining_quantity': str(round(remaining_quantity, 2)),
                    'remaining_price': str(round(remaining_price, 4)),
                    'remaining_price_total': str(round(remaining_price_total, 2))
                }
                kardex_dict.append(item)

                last_remaining_quantity = remaining_quantity
                last_remaining_price_total = remaining_price_total

        else:
            previous_movements = Kardex.objects.filter(
                product=product_obj,
                create_at__lt=start_date
            ).order_by('-create_at')

            last_previous_movement = previous_movements.first()
            if last_previous_movement:
                last_month_id = last_previous_movement.id
                last_month_remaining_quantity = last_previous_movement.remaining_quantity
                last_month_remaining_price_total = last_previous_movement.remaining_price_total
            else:
                last_month_id = None
                last_month_remaining_quantity = 0
                last_month_remaining_price_total = 0

            try:
                if last_month_remaining_quantity == 0:
                    last_average_cost = decimal.Decimal('0')
                else:
                    last_average_cost = last_month_remaining_price_total / last_month_remaining_quantity
            except DivisionUndefined:
                last_average_cost = decimal.Decimal('0')
            remaining_quantity = decimal.Decimal(last_month_remaining_quantity)
            remaining_price_total = decimal.Decimal(last_month_remaining_price_total)
            first_date = datetime.strptime(date_product + '-01', "%Y-%m-%d")

            item = {
                'id': last_month_id,
                'operation': 'C',
                'date': first_date.strftime("%d/%m/%Y"),
                'type_document': '00',
                'serial': 'INICIAL',
                'number': '0',
                'detail': 'SALDO INICIAL',
                'type_operation': '16',
                'quantity': str(round(last_month_remaining_quantity, 2)),
                'unit_cost': '',
                'total_cost': str(round(last_month_remaining_price_total, 2)),
                'remaining_quantity': str(round(last_month_remaining_quantity, 2)),
                'remaining_price': str(round(last_average_cost, 4)),
                'remaining_price_total': str(round(last_month_remaining_price_total, 2))
            }
            kardex_dict.append(item)

            last_remaining_quantity = remaining_quantity
            last_remaining_price_total = remaining_price_total

        sell_unit = (last_month_remaining_quantity + sum_quantities_entries) - last_remaining_quantity
        cost_sale = (last_month_remaining_price_total + sum_total_cost_entries) - last_remaining_price_total

        tpl = loader.get_template('sales/new_kardex_grid_list.html')
        context = ({
            'kardex_dict': kardex_dict,
            'sum_quantities_entries': sum_quantities_entries,
            'sum_total_cost_entries': str(round(sum_total_cost_entries, 2)),
            'sum_quantities_exits': sum_quantities_exits,
            'sum_total_cost_exits': str(round(sum_total_cost_exits, 2)),
            'sum_remaining_price': round(sum_remaining_price, 4),
            'last_month_remaining_quantity': str(round(last_month_remaining_quantity, 2)),
            'last_month_remaining_price_total': str(round(last_month_remaining_price_total, 2)),
            'last_remaining_quantity': str(round(last_remaining_quantity, 2)),
            'last_remaining_price_total': str(round(last_remaining_price_total, 2)),
            'sell_unit': str(round(sell_unit, 2)),
            'cost_sale': str(round(cost_sale, 2)),
            'product': product_obj,
            'formatdate': formatdate
        })
        return JsonResponse({
            'success': True,
            'grid': tpl.render(context, request),
            'product_name': product_obj.complete_name(),
        }, status=HTTPStatus.OK)

        # else:
        #     return JsonResponse({
        #         'success': False,
        #         'message': 'El Producto {}, no cuenta con registros en el mes selecccionado'.format(
        #             product_obj.complete_name())
        #     }, status=HTTPStatus.OK)


def recalculate_credit_note(request):
    if request.method == 'GET':
        order_set = Order.objects.filter(status='N', add='N', type='V').order_by('id')
        for o in order_set:
            note_qr = o.note_qr
            note_parts = note_qr.split('|')
            if len(note_parts) > 0:
                value_str = note_parts[5]
                value_num = float(value_str.replace(',', ''))
                o.note_total = decimal.Decimal(value_num)
                o.save()
        return JsonResponse({
            'success': True,
        }, status=HTTPStatus.OK)


def update_invoice_stock_product(request):
    if request.method == 'GET':
        products_set = Product.objects.all()
        for p in products_set:
            kardex_set = Kardex.objects.filter(product=p)
            if kardex_set.exists():
                kardex_obj = kardex_set.last()
                p.invoice_stock = kardex_obj.remaining_quantity
                p.save()

        return JsonResponse({
            'success': True,
        }, status=HTTPStatus.OK)


def fill_kardex_invoice(request):
    if request.method == 'GET':
        date_initial = '2022-11-01'
        date_final = '2022-12-31'
        # products_set = Product.objects.all()
        products_set = Product.objects.filter(id=5403)
        for p in products_set:
            # sales_details = OrderDetail.objects.filter(Q(product=p,
            #                                              create_at__range=(date_initial, date_final),
            #                                              order__type='V', order__doc__in=['1', '2'],
            #                                              order__bill_type__isnull=False,
            #                                              order__bill_number__isnull=False,
            #                                              order__status__in=['E', 'N']) |
            #                                            Q(product=p, order__number__isnull=False,
            #                                              create_at__range=(date_initial, date_final),
            #                                              order__type='C')).distinct('order_id').order_by('create_at')

            # sales_details2 = OrderDetail.objects.filter(
            #         order__number__isnull=False,
            #         order__invoice_date__range=(date_initial, date_final),
            #         order__type='C',
            #         order__status='R'
            # ).annotate(
            #     min_create_at=Min('create_at')
            # ).order_by('min_create_at', 'id')

            # sales_details2 = OrderDetail.objects.filter(
            #     Q(product=p,
            #       create_at__range=(date_initial, date_final),
            #       order__type='V',
            #       order__doc__in=['1', '2'],
            #       order__bill_type__isnull=False,
            #       order__bill_number__isnull=False,
            #       order__status__in=['E', 'N']) |
            #     Q(product=p,
            #       order__number__isnull=False,
            #       create_at__range=(date_initial, date_final),
            #       order__type='C')
            # ).annotate(
            #     min_create_at=Min('create_at')
            # ).order_by('min_create_at', 'id')

            sales_details = OrderDetail.objects.filter(
                Q(product=p,
                  create_at__range=(date_initial, date_final),
                  order__type='V',
                  order__doc__in=['1', '2'],
                  order__bill_type__isnull=False,
                  order__bill_number__isnull=False,
                  order__status__in=['E', 'N']) |
                Q(product=p,
                  order__number__isnull=False,
                  order__invoice_date__range=(date_initial, date_final),
                  order__type='C')
            ).annotate(
                min_create_at=Min(
                    Case(
                        When(order__type='V', then='create_at'),
                        When(order__type='C', then='order__invoice_date'),
                        default=Value(None),
                        output_field=models.DateTimeField(),
                    )
                )
            ).order_by('min_create_at', 'id')

            for s in sales_details:

                price = s.price

                kardex_set = Kardex.objects.filter(product=p)
                quantity = s.quantity
                quantity_in_unit = 0

                create_date = s.create_at

                if s.order.type == 'C':
                    create_date = s.order.invoice_date
                    if s.order.coin == '2':
                        price = s.price * s.order.change

                amount = round(s.quantity * price, 6)

                if s.unit == 'KGM':
                    quantity_in_unit = round(decimal.Decimal(s.quantity_niu), 4)
                    # amount = round(decimal.Decimal(s.quantity) * s.price, 6)
                    # quantity = round(decimal.Decimal(s.quantity) * decimal.Decimal(s.quantity_niu), 4)
                    amount = round(quantity * s.price, 6)

                if kardex_set.exists():
                    type_document = '00'
                    if s.order.doc == '1':
                        type_document = str('0' + s.order.doc)
                    elif s.order.doc == '2':
                        type_document = '03'

                    if s.order.type == 'C' and s.order.status == 'R':

                        if s.order.coin == '2':
                            sum_total_purchase = decimal.Decimal(round(s.order.sum_total_purchase() * s.order.change, 2))
                            total_purchase = decimal.Decimal(round(s.order.total * s.order.change, 2))
                            if sum_total_purchase == total_purchase:
                                amount = decimal.Decimal(quantity) * decimal.Decimal(price) / decimal.Decimal(1.18)
                        else:
                            sum_total_purchase = s.order.sum_total_purchase()
                            total_purchase = s.order.total
                            if sum_total_purchase == total_purchase:
                                amount = s.amount_without_igv()

                        kardex_input(product=p, quantity=quantity, total_cost=amount,
                                     order_detail_obj=s, type_document=type_document, type_operation='02',
                                     create_at=create_date, quantity_in_unit=quantity_in_unit)
                    elif s.order.type == 'C' and (s.order.status == 'A' or s.order.status == 'N'):

                        kardex_ouput(product=p, quantity=quantity, order_detail_obj=s,
                                     type_document=str('0' + s.order.doc), type_operation='05', create_at=create_date,
                                     quantity_in_unit=quantity_in_unit)

                    elif s.order.type == 'V' and s.order.status == 'E':
                        kardex_ouput(product=p, quantity=quantity, order_detail_obj=s,
                                     type_document=type_document, type_operation='01', create_at=create_date,
                                     quantity_in_unit=quantity_in_unit)

                    elif s.order.type == 'V' and s.order.status == 'N':
                        kardex_input(product=p, quantity=quantity, total_cost=amount, order_detail_obj=s,
                                     type_document='07', type_operation='06', create_at=create_date,
                                     quantity_in_unit=quantity_in_unit)
                else:
                    if s.order.type == 'C' and s.order.status == 'R':

                        sum_total_purchase = s.order.sum_total_purchase()
                        total_purchase = s.order.total
                        if decimal.Decimal(round(sum_total_purchase, 2)) == decimal.Decimal(round(total_purchase, 2)):
                            price = decimal.Decimal(price) / decimal.Decimal(1.18)

                        kardex_initial(product=p, stock=quantity, price_unit=price,
                                       order_detail_obj=s, create_at=create_date, quantity_in_unit=quantity_in_unit)
                    else:
                        kardex_initial(product=p, stock=quantity, price_unit=price,
                                       order_detail_obj=s, create_at=create_date, quantity_in_unit=quantity_in_unit)

        return JsonResponse({
            'success': True,
        }, status=HTTPStatus.OK)


def test_socket(request):
    if request.method == 'GET':
        return render(request, 'sales/test_socket.html', {
        })