import datetime
import decimal
import pandas as pd
import calendar

from django.db.models.functions import Coalesce

from unidecode import unidecode
from django.db.models import F, Prefetch, OuterRef, Subquery, Q, Sum, Value, DecimalField
from django.shortcuts import render
from django.http.response import HttpResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from apps.sales.models import Product, OrderDetail, Order, Presentation, Kardex


class ReportProduct(TemplateView):
    # @method_decorator(csrf_exempt)
    # def dispatch(self, request, *args, **kwargs):
    #     return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # campo = int(request.GET.get('campo'))
        query = Product.objects.all()
        wb = Workbook()
        bandera = True
        cont = 1
        row = 4
        if bandera:
            ws = wb.active
            ws.title = 'LISTADO DE PRODUCTOS' + str(cont)  # hoja
            bandera = False
        else:
            ws = wb.create_sheet('LISTADO DE PRODUCTOS' + str(cont))
        # Crear el título en la hoja
        my_date = datetime.datetime.now()
        date_now = my_date.strftime("%d-%m-%Y")
        hour_now = my_date.strftime("%H:%M:%S")
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B1'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
        ws['B1'] = 'REPORTE DE PRODUCTOS DE PERNOS JHUNIOR' + ' FECHA: ' + str(date_now) + ' HORA: ' + str(hour_now)

        # Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:N1')

        ws.row_dimensions[1].height = 25

        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 9
        ws.column_dimensions['F'].width = 9
        ws.column_dimensions['G'].width = 9
        ws.column_dimensions['H'].width = 9
        ws.column_dimensions['I'].width = 9
        ws.column_dimensions['J'].width = 9
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 9
        ws.column_dimensions['M'].width = 9
        ws.column_dimensions['N'].width = 25

        # Crear la cabecera
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['B3'] = 'CODIGO'

        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['C3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['C3'] = 'NOMBRE PRODUCTO'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['D3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['D3'] = 'DESCRIPCION PRODUCTO'

        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['E3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['E3'] = 'MARCA'

        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['F3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['F3'] = 'FAMILIA'

        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['G3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['G3'] = 'TIPO'

        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['H3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['H3'] = 'ANCHO'

        ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['I3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['I3'] = 'LARGO'

        ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['J3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['J3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['J3'] = 'ALTO'

        ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['K3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['K3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['K3'] = 'ALMACEN'

        ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['L3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['L3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['L3'] = 'STOCK'

        ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['M3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['M3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['M3'] = 'MINIMO'

        ws['N3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['N3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['N3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['N3'] = 'PRODUCTO RELACIONADO'

        for p in query.order_by('id'):
            # Pintamos los datos en el reporte
            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
            text_color = Font(name='Calibri', size=9)
            border_cell = Border(left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 top=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))
            align_cell = Alignment(horizontal="center")
            if p.stock <= p.minimum:
                color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                text_color = Font(name='Calibri', size=9, color='00FF0000')
            ws.cell(row=row, column=2).alignment = align_cell
            ws.cell(row=row, column=2).border = border_cell
            ws.cell(row=row, column=2).fill = color_cell
            ws.cell(row=row, column=2).font = text_color
            ws.cell(row=row, column=2).value = p.code

            ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=3).border = border_cell
            ws.cell(row=row, column=3).fill = color_cell
            ws.cell(row=row, column=3).font = text_color
            ws.cell(row=row, column=3).value = p.name

            ws.cell(row=row, column=4).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=4).border = border_cell
            ws.cell(row=row, column=4).fill = color_cell
            ws.cell(row=row, column=4).font = text_color
            ws.cell(row=row, column=4).value = p.description

            ws.cell(row=row, column=5).alignment = align_cell
            ws.cell(row=row, column=5).border = border_cell
            ws.cell(row=row, column=5).fill = color_cell
            ws.cell(row=row, column=5).font = text_color
            ws.cell(row=row, column=5).value = p.brand.name

            ws.cell(row=row, column=6).alignment = align_cell
            ws.cell(row=row, column=6).border = border_cell
            ws.cell(row=row, column=6).fill = color_cell
            ws.cell(row=row, column=6).font = text_color
            ws.cell(row=row, column=6).value = p.family.name

            ws.cell(row=row, column=7).alignment = align_cell
            ws.cell(row=row, column=7).border = border_cell
            ws.cell(row=row, column=7).fill = color_cell
            ws.cell(row=row, column=7).font = text_color
            ws.cell(row=row, column=7).value = p.get_type_display()

            ws.cell(row=row, column=8).alignment = align_cell
            ws.cell(row=row, column=8).border = border_cell
            ws.cell(row=row, column=8).fill = color_cell
            ws.cell(row=row, column=8).font = text_color
            ws.cell(row=row, column=8).value = p.width

            ws.cell(row=row, column=9).alignment = align_cell
            ws.cell(row=row, column=9).border = border_cell
            ws.cell(row=row, column=9).fill = color_cell
            ws.cell(row=row, column=9).font = text_color
            ws.cell(row=row, column=9).value = p.length

            ws.cell(row=row, column=10).alignment = align_cell
            ws.cell(row=row, column=10).border = border_cell
            ws.cell(row=row, column=10).fill = color_cell
            ws.cell(row=row, column=10).font = text_color
            ws.cell(row=row, column=10).value = p.height

            ws.cell(row=row, column=11).alignment = align_cell
            ws.cell(row=row, column=11).border = border_cell
            ws.cell(row=row, column=11).fill = color_cell
            ws.cell(row=row, column=11).font = text_color
            ws.cell(row=row, column=11).value = p.store

            ws.cell(row=row, column=12).alignment = align_cell
            ws.cell(row=row, column=12).border = border_cell
            ws.cell(row=row, column=12).fill = color_cell
            ws.cell(row=row, column=12).font = text_color
            ws.cell(row=row, column=12).value = p.stock

            ws.cell(row=row, column=13).alignment = align_cell
            ws.cell(row=row, column=13).border = border_cell
            ws.cell(row=row, column=13).fill = color_cell
            ws.cell(row=row, column=13).font = text_color
            ws.cell(row=row, column=13).value = p.minimum

            ws.cell(row=row, column=14).alignment = align_cell
            ws.cell(row=row, column=14).border = border_cell
            ws.cell(row=row, column=14).fill = color_cell
            ws.cell(row=row, column=14).font = text_color
            ws.cell(row=row, column=14).value = p.relation

            cont += 1
            row += 1

        # Establecer el nombre de mi archivo
        nombre_archivo = "Productos.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class StockMin(TemplateView):

    def get(self, request, *args, **kwargs):
        query = Product.objects.filter(stock__lte=F('minimum'))
        wb = Workbook()
        bandera = True
        cont = 1
        row = 4
        if bandera:
            ws = wb.active
            ws.title = 'STOCK DE PRODUCTOS' + str(cont)  # hoja
            bandera = False
        else:
            ws = wb.create_sheet('STOCK DE PRODUCTOS' + str(cont))
        # Crear el título en la hoja
        my_date = datetime.datetime.now()
        date_now = my_date.strftime("%d-%m-%Y")
        hour_now = my_date.strftime("%H:%M:%S")
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type="solid")
        ws['B1'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
        ws['B1'] = 'PRODUCTOS CON STOCK AL MINIMO ' + ' FECHA: ' + str(date_now) + ' HORA: ' + str(hour_now)

        # Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:N1')

        ws.row_dimensions[1].height = 25

        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 9
        ws.column_dimensions['F'].width = 9
        ws.column_dimensions['G'].width = 9
        ws.column_dimensions['H'].width = 9
        ws.column_dimensions['I'].width = 9
        ws.column_dimensions['J'].width = 9
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 9
        ws.column_dimensions['M'].width = 9
        ws.column_dimensions['N'].width = 25

        # Crear la cabecera
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['B3'] = 'CODIGO'

        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['C3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['C3'] = 'NOMBRE PRODUCTO'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['D3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['D3'] = 'DESCRIPCION PRODUCTO'

        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['E3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['E3'] = 'MARCA'

        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['F3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['F3'] = 'FAMILIA'

        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['G3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['G3'] = 'TIPO'

        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['H3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['H3'] = 'ANCHO'

        ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['I3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['I3'] = 'LARGO'

        ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['J3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['J3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['J3'] = 'ALTO'

        ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['K3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['K3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['K3'] = 'ALMACEN'

        ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['L3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['L3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['L3'] = 'STOCK'

        ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['M3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['M3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['M3'] = 'MINIMO'

        ws['N3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['N3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['N3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['N3'] = 'PRODUCTO RELACIONADO'

        for p in query.order_by('id'):
            # Pintamos los datos en el reporte
            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
            stock_cell = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type="solid")
            text_color = Font(name='Calibri', size=9)
            border_cell = Border(left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 top=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))
            align_cell = Alignment(horizontal="center")
            ws.cell(row=row, column=2).alignment = align_cell
            ws.cell(row=row, column=2).border = border_cell
            ws.cell(row=row, column=2).fill = color_cell
            ws.cell(row=row, column=2).font = text_color
            ws.cell(row=row, column=2).value = p.code

            ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=3).border = border_cell
            ws.cell(row=row, column=3).fill = color_cell
            ws.cell(row=row, column=3).font = text_color
            ws.cell(row=row, column=3).value = p.name

            ws.cell(row=row, column=4).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=4).border = border_cell
            ws.cell(row=row, column=4).fill = color_cell
            ws.cell(row=row, column=4).font = text_color
            ws.cell(row=row, column=4).value = p.description

            ws.cell(row=row, column=5).alignment = align_cell
            ws.cell(row=row, column=5).border = border_cell
            ws.cell(row=row, column=5).fill = color_cell
            ws.cell(row=row, column=5).font = text_color
            ws.cell(row=row, column=5).value = p.brand.name

            ws.cell(row=row, column=6).alignment = align_cell
            ws.cell(row=row, column=6).border = border_cell
            ws.cell(row=row, column=6).fill = color_cell
            ws.cell(row=row, column=6).font = text_color
            ws.cell(row=row, column=6).value = p.family.name

            ws.cell(row=row, column=7).alignment = align_cell
            ws.cell(row=row, column=7).border = border_cell
            ws.cell(row=row, column=7).fill = color_cell
            ws.cell(row=row, column=7).font = text_color
            ws.cell(row=row, column=7).value = p.get_type_display()

            ws.cell(row=row, column=8).alignment = align_cell
            ws.cell(row=row, column=8).border = border_cell
            ws.cell(row=row, column=8).fill = color_cell
            ws.cell(row=row, column=8).font = text_color
            ws.cell(row=row, column=8).value = p.width

            ws.cell(row=row, column=9).alignment = align_cell
            ws.cell(row=row, column=9).border = border_cell
            ws.cell(row=row, column=9).fill = color_cell
            ws.cell(row=row, column=9).font = text_color
            ws.cell(row=row, column=9).value = p.length

            ws.cell(row=row, column=10).alignment = align_cell
            ws.cell(row=row, column=10).border = border_cell
            ws.cell(row=row, column=10).fill = color_cell
            ws.cell(row=row, column=10).font = text_color
            ws.cell(row=row, column=10).value = p.height

            ws.cell(row=row, column=11).alignment = align_cell
            ws.cell(row=row, column=11).border = border_cell
            ws.cell(row=row, column=11).fill = color_cell
            ws.cell(row=row, column=11).font = text_color
            ws.cell(row=row, column=11).value = p.store

            ws.cell(row=row, column=12).alignment = align_cell
            ws.cell(row=row, column=12).border = border_cell
            ws.cell(row=row, column=12).fill = stock_cell
            ws.cell(row=row, column=12).font = text_color
            ws.cell(row=row, column=12).value = p.stock

            ws.cell(row=row, column=13).alignment = align_cell
            ws.cell(row=row, column=13).border = border_cell
            ws.cell(row=row, column=13).fill = color_cell
            ws.cell(row=row, column=13).font = text_color
            ws.cell(row=row, column=13).value = p.minimum

            ws.cell(row=row, column=14).alignment = align_cell
            ws.cell(row=row, column=14).border = border_cell
            ws.cell(row=row, column=14).fill = color_cell
            ws.cell(row=row, column=14).font = text_color
            ws.cell(row=row, column=14).value = p.relation

            cont += 1
            row += 1

        # Establecer el nombre de mi archivo
        nombre_archivo = "MinStock.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


def reportkardex(request, init=None, end=None, pk=None):
    if init and end and pk:
        product_obj = Product.objects.get(id=int(pk))
        # m = month[5:]
        detail_set = OrderDetail.objects.filter(product=product_obj, order__type__in=['V', 'C'],
                                                order__create_at__range=(init, end)).order_by('id')

        order_detail_obj = OrderDetail.objects.filter(product=product_obj, order__type__in=['V', 'C']).first()
        wb = Workbook()
        bandera = True
        cont = 1
        row = 4
        if bandera:
            ws = wb.active
            ws.title = str("PRODUCTO")  # hoja
            bandera = False
        else:
            ws = wb.create_sheet(str(product_obj.name) + str(cont))
        # Crear el título en la hoja
        my_date = datetime.datetime.now()
        date_now = my_date.strftime("%d-%m-%Y")
        hour_now = my_date.strftime("%H:%M:%S")
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B1'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
        ws['B1'] = str(product_obj.name.upper())

        # Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:J1')

        ws.row_dimensions[1].height = 25

        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 15

        # Crear la cabecera
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['B3'] = 'FECHA'

        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['C3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['C3'] = 'ORDEN'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['D3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['D3'] = 'OPERACIÓN'

        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['E3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['E3'] = 'CANTIDAD'

        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['F3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['F3'] = 'MEDIDA'

        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['G3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['G3'] = 'UNIDADES'

        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['H3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['H3'] = 'PRECIO'

        ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['I3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['I3'] = 'TOTAL'

        ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['J3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['J3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['J3'] = 'RESTANTE'
        counter_for = 0
        for d in detail_set:
            if counter_for == 0 and order_detail_obj.id == d.id:
                color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                text_color = Font(name='Calibri', size=9)
                border_cell = Border(left=Side(border_style="thin"),
                                     right=Side(border_style="thin"),
                                     top=Side(border_style="thin"),
                                     bottom=Side(border_style="thin"))
                align_cell = Alignment(horizontal="center")

                ws.cell(row=row, column=2).alignment = align_cell
                ws.cell(row=row, column=2).border = border_cell
                ws.cell(row=row, column=2).fill = color_cell
                ws.cell(row=row, column=2).font = text_color
                ws.cell(row=row, column=2).value = "-"

                ws.cell(row=row, column=3).alignment = align_cell
                ws.cell(row=row, column=3).border = border_cell
                ws.cell(row=row, column=3).fill = color_cell
                ws.cell(row=row, column=3).font = text_color
                ws.cell(row=row, column=3).value = "INVENTARIO"

                ws.cell(row=row, column=4).alignment = align_cell
                ws.cell(row=row, column=4).border = border_cell
                ws.cell(row=row, column=4).fill = color_cell
                ws.cell(row=row, column=4).font = text_color
                ws.cell(row=row, column=4).value = "ENTRADA"

                ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
                ws.cell(row=row, column=5).border = border_cell
                ws.cell(row=row, column=5).fill = color_cell
                ws.cell(row=row, column=5).font = text_color
                ws.cell(row=row, column=5).value = round(d.previous(), 2)

                ws.cell(row=row, column=6).alignment = align_cell
                ws.cell(row=row, column=6).border = border_cell
                ws.cell(row=row, column=6).fill = color_cell
                ws.cell(row=row, column=6).font = text_color
                ws.cell(row=row, column=6).value = "UNIDAD"

                ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
                ws.cell(row=row, column=7).border = border_cell
                ws.cell(row=row, column=7).fill = color_cell
                ws.cell(row=row, column=7).font = text_color
                ws.cell(row=row, column=7).value = round(d.previous(), 2)

                ws.cell(row=row, column=8).alignment = Alignment(horizontal="right")
                ws.cell(row=row, column=8).border = border_cell
                ws.cell(row=row, column=8).fill = color_cell
                ws.cell(row=row, column=8).font = text_color
                ws.cell(row=row, column=8).value = "-"

                ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
                ws.cell(row=row, column=9).border = border_cell
                ws.cell(row=row, column=9).fill = color_cell
                ws.cell(row=row, column=9).font = text_color
                ws.cell(row=row, column=9).value = "-"

                ws.cell(row=row, column=10).alignment = Alignment(horizontal="right")
                ws.cell(row=row, column=10).border = border_cell
                ws.cell(row=row, column=10).fill = color_cell
                ws.cell(row=row, column=10).font = text_color
                ws.cell(row=row, column=10).value = round(d.previous(), 2)
                row += 1
            # Pintamos los datos en el reporte
            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
            text_color = Font(name='Calibri', size=9)
            border_cell = Border(left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 top=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))
            align_cell = Alignment(horizontal="center")

            ws.cell(row=row, column=2).alignment = align_cell
            ws.cell(row=row, column=2).border = border_cell
            ws.cell(row=row, column=2).fill = color_cell
            ws.cell(row=row, column=2).font = text_color
            ws.cell(row=row, column=2).value = d.order.create_at.strftime("%d-%m-%Y")

            ws.cell(row=row, column=3).alignment = align_cell
            ws.cell(row=row, column=3).border = border_cell
            ws.cell(row=row, column=3).fill = color_cell
            ws.cell(row=row, column=3).font = text_color
            ws.cell(row=row, column=3).value = d.order.get_type_display()

            ws.cell(row=row, column=4).alignment = align_cell
            ws.cell(row=row, column=4).border = border_cell
            ws.cell(row=row, column=4).fill = color_cell
            ws.cell(row=row, column=4).font = text_color
            ws.cell(row=row, column=4).value = str(d.get_operation_display()).upper()

            ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=5).border = border_cell
            ws.cell(row=row, column=5).fill = color_cell
            ws.cell(row=row, column=5).font = text_color
            ws.cell(row=row, column=5).value = round(d.quantity, 2)

            ws.cell(row=row, column=6).alignment = align_cell
            ws.cell(row=row, column=6).border = border_cell
            ws.cell(row=row, column=6).fill = color_cell
            ws.cell(row=row, column=6).font = text_color
            ws.cell(row=row, column=6).value = d.get_unit_display()

            ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=7).border = border_cell
            ws.cell(row=row, column=7).fill = color_cell
            ws.cell(row=row, column=7).font = text_color
            ws.cell(row=row, column=7).value = d.quantity_niu

            ws.cell(row=row, column=8).alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=8).border = border_cell
            ws.cell(row=row, column=8).fill = color_cell
            ws.cell(row=row, column=8).font = text_color
            ws.cell(row=row, column=8).value = round(d.price, 4)

            ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=9).border = border_cell
            ws.cell(row=row, column=9).fill = color_cell
            ws.cell(row=row, column=9).font = text_color
            ws.cell(row=row, column=9).value = round(d.amount(), 4)

            ws.cell(row=row, column=10).alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=10).border = border_cell
            ws.cell(row=row, column=10).fill = color_cell
            ws.cell(row=row, column=10).font = text_color
            ws.cell(row=row, column=10).value = round(d.quantity_remaining, 2)
            counter_for += 1
            cont += 1
            row += 1

        # Establecer el nombre de mi archivo
        nombre_archivo = "Kardex.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class FilterProduct(TemplateView):

    def get(self, request, *args, **kwargs):
        query = Product.objects.filter(stock__lte=F('minimum'))
        detail = request.GET
        wb = Workbook()
        bandera = True
        cont = 1
        row = 4
        if bandera:
            ws = wb.active
            ws.title = 'STOCK DE PRODUCTOS' + str(cont)  # hoja
            bandera = False
        else:
            ws = wb.create_sheet('STOCK DE PRODUCTOS' + str(cont))
        # Crear el título en la hoja
        my_date = datetime.datetime.now()
        date_now = my_date.strftime("%d-%m-%Y")
        hour_now = my_date.strftime("%H:%M:%S")
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type="solid")
        ws['B1'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
        ws['B1'] = 'PRODUCTOS CON STOCK AL MINIMO ' + ' FECHA: ' + str(date_now) + ' HORA: ' + str(hour_now)

        # Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:N1')

        ws.row_dimensions[1].height = 25

        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 9
        ws.column_dimensions['F'].width = 9
        ws.column_dimensions['G'].width = 9
        ws.column_dimensions['H'].width = 9
        ws.column_dimensions['I'].width = 9
        ws.column_dimensions['J'].width = 9
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 9
        ws.column_dimensions['M'].width = 9
        ws.column_dimensions['N'].width = 25

        # Crear la cabecera
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['B3'] = 'CODIGO'

        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['C3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['C3'] = 'NOMBRE PRODUCTO'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['D3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['D3'] = 'DESCRIPCION PRODUCTO'

        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['E3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['E3'] = 'MARCA'

        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['F3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['F3'] = 'FAMILIA'

        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['G3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['G3'] = 'TIPO'

        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['H3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['H3'] = 'ANCHO'

        ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['I3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['I3'] = 'LARGO'

        ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['J3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['J3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['J3'] = 'ALTO'

        ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['K3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['K3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['K3'] = 'ALMACEN'

        ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['L3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['L3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['L3'] = 'STOCK'

        ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['M3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['M3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['M3'] = 'MINIMO'

        ws['N3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['N3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['N3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['N3'] = 'PRODUCTO RELACIONADO'

        for p in query.order_by('id'):
            # Pintamos los datos en el reporte
            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
            stock_cell = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type="solid")
            text_color = Font(name='Calibri', size=9)
            border_cell = Border(left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 top=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))
            align_cell = Alignment(horizontal="center")
            ws.cell(row=row, column=2).alignment = align_cell
            ws.cell(row=row, column=2).border = border_cell
            ws.cell(row=row, column=2).fill = color_cell
            ws.cell(row=row, column=2).font = text_color
            ws.cell(row=row, column=2).value = p.code

            ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=3).border = border_cell
            ws.cell(row=row, column=3).fill = color_cell
            ws.cell(row=row, column=3).font = text_color
            ws.cell(row=row, column=3).value = p.name

            ws.cell(row=row, column=4).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=4).border = border_cell
            ws.cell(row=row, column=4).fill = color_cell
            ws.cell(row=row, column=4).font = text_color
            ws.cell(row=row, column=4).value = p.description

            ws.cell(row=row, column=5).alignment = align_cell
            ws.cell(row=row, column=5).border = border_cell
            ws.cell(row=row, column=5).fill = color_cell
            ws.cell(row=row, column=5).font = text_color
            ws.cell(row=row, column=5).value = p.brand.name

            ws.cell(row=row, column=6).alignment = align_cell
            ws.cell(row=row, column=6).border = border_cell
            ws.cell(row=row, column=6).fill = color_cell
            ws.cell(row=row, column=6).font = text_color
            ws.cell(row=row, column=6).value = p.family.name

            ws.cell(row=row, column=7).alignment = align_cell
            ws.cell(row=row, column=7).border = border_cell
            ws.cell(row=row, column=7).fill = color_cell
            ws.cell(row=row, column=7).font = text_color
            ws.cell(row=row, column=7).value = p.get_type_display()

            ws.cell(row=row, column=8).alignment = align_cell
            ws.cell(row=row, column=8).border = border_cell
            ws.cell(row=row, column=8).fill = color_cell
            ws.cell(row=row, column=8).font = text_color
            ws.cell(row=row, column=8).value = p.width

            ws.cell(row=row, column=9).alignment = align_cell
            ws.cell(row=row, column=9).border = border_cell
            ws.cell(row=row, column=9).fill = color_cell
            ws.cell(row=row, column=9).font = text_color
            ws.cell(row=row, column=9).value = p.length

            ws.cell(row=row, column=10).alignment = align_cell
            ws.cell(row=row, column=10).border = border_cell
            ws.cell(row=row, column=10).fill = color_cell
            ws.cell(row=row, column=10).font = text_color
            ws.cell(row=row, column=10).value = p.height

            ws.cell(row=row, column=11).alignment = align_cell
            ws.cell(row=row, column=11).border = border_cell
            ws.cell(row=row, column=11).fill = color_cell
            ws.cell(row=row, column=11).font = text_color
            ws.cell(row=row, column=11).value = p.store

            ws.cell(row=row, column=12).alignment = align_cell
            ws.cell(row=row, column=12).border = border_cell
            ws.cell(row=row, column=12).fill = stock_cell
            ws.cell(row=row, column=12).font = text_color
            ws.cell(row=row, column=12).value = p.stock

            ws.cell(row=row, column=13).alignment = align_cell
            ws.cell(row=row, column=13).border = border_cell
            ws.cell(row=row, column=13).fill = color_cell
            ws.cell(row=row, column=13).font = text_color
            ws.cell(row=row, column=13).value = p.minimum

            ws.cell(row=row, column=14).alignment = align_cell
            ws.cell(row=row, column=14).border = border_cell
            ws.cell(row=row, column=14).fill = color_cell
            ws.cell(row=row, column=14).font = text_color
            ws.cell(row=row, column=14).value = p.relation

            cont += 1
            row += 1

        # Establecer el nombre de mi archivo
        nombre_archivo = "MinStock.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


# class ReportProduct(TemplateView):
#     def get(self, request, *args, **kwargs):
#         # campo = int(request.GET.get('campo'))
#         query = Product.objects.filter(id__lte=10)
#         wb = Workbook()
#         bandera = True
#         cont = 1
#         controlador = 4
#         for q in query:
#             if bandera:
#                 ws = wb.active
#                 ws.title = 'Hoja' + str(cont)
#                 bandera = False
#             else:
#                 ws = wb.create_sheet('Hoja' + str(cont))
#             # Crear el título en la hoja
#             ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
#             ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                      top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#             ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type="solid")
#             ws['B1'].font = Font(name='Calibri', size=12, bold=True)
#             ws['B1'] = 'REPORTE PERSONALIZADO EN EXCEL CON DJANGO'
#
#             # Cambiar caracteristicas de las celdas
#             ws.merge_cells('B1:E1')
#
#             ws.row_dimensions[1].height = 25
#
#             ws.column_dimensions['B'].width = 20
#             ws.column_dimensions['C'].width = 20
#             ws.column_dimensions['D'].width = 20
#             ws.column_dimensions['E'].width = 20
#
#             # Crear la cabecera
#             ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
#             ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                      top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#             ws['B3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
#             ws['B3'].font = Font(name='Calibro', size=10, bold=True)
#             ws['B3'] = 'Nombres'
#
#             ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
#             ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                      top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#             ws['C3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
#             ws['C3'].font = Font(name='Calibro', size=10, bold=True)
#             ws['C3'] = 'Apellidos'
#
#             ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
#             ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                      top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#             ws['D3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
#             ws['D3'].font = Font(name='Calibro', size=10, bold=True)
#             ws['D3'] = 'Dirección'
#
#             ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
#             ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
#                                      top=Side(border_style="thin"), bottom=Side(border_style="thin"))
#             ws['E3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
#             ws['E3'].font = Font(name='Calibro', size=10, bold=True)
#             ws['E3'] = 'Edad'
#
#             # Pintamos los datos en el reporte
#             ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
#             ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
#                                                                right=Side(border_style="thin"),
#                                                                top=Side(border_style="thin"),
#                                                                bottom=Side(border_style="thin"))
#             ws.cell(row=controlador, column=2).font = Font(name='Calibri', size=8)
#             ws.cell(row=controlador, column=2).value = q.name
#
#             ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
#             ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
#                                                                right=Side(border_style="thin"),
#                                                                top=Side(border_style="thin"),
#                                                                bottom=Side(border_style="thin"))
#             ws.cell(row=controlador, column=3).font = Font(name='Calibri', size=8)
#             ws.cell(row=controlador, column=3).value = q.code
#
#             ws.cell(row=controlador, column=4).alignment = Alignment(horizontal="center")
#             ws.cell(row=controlador, column=4).border = Border(left=Side(border_style="thin"),
#                                                                right=Side(border_style="thin"),
#                                                                top=Side(border_style="thin"),
#                                                                bottom=Side(border_style="thin"))
#             ws.cell(row=controlador, column=4).font = Font(name='Calibri', size=8)
#             ws.cell(row=controlador, column=4).value = q.brand.name
#
#             ws.cell(row=controlador, column=5).alignment = Alignment(horizontal="center")
#             ws.cell(row=controlador, column=5).border = Border(left=Side(border_style="thin"),
#                                                                right=Side(border_style="thin"),
#                                                                top=Side(border_style="thin"),
#                                                                bottom=Side(border_style="thin"))
#             ws.cell(row=controlador, column=5).font = Font(name='Calibri', size=8)
#             ws.cell(row=controlador, column=5).value = q.family.name
#
#             cont += 1
#
#         # Establecer el nombre de mi archivo
#         nombre_archivo = "Productos.xlsx"
#         # Definir el tipo de respuesta que se va a dar
#         response = HttpResponse(content_type="application/ms-excel")
#         contenido = "attachment; filename = {0}".format(nombre_archivo)
#         response["Content-Disposition"] = contenido
#         wb.save(response)
#         return response
def purchase_excel(request, init=None, end=None):
    if init and end:
        order_set = Order.objects.filter(create_at__range=(init, end), type='C', status__in=['R', 'E', 'P'],
                                         total__gt=0).order_by('id')
        wb = Workbook()
        bandera = True
        cont = 1
        row = 4
        if bandera:
            ws = wb.active
            ws.title = str("Reporte-compras")  # hoja
            bandera = False
        else:
            ws = wb.create_sheet(str("Reporte-compras") + str(cont))
        # Crear el título en la hoja
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B1'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
        ws['B1'] = str("REPORTE DE COMPRAS")
        # " " + str(datetime.date(init).strptime("%d-%m-%Y")) + "-" + str(datetime.date(end).strftime("%d-%m-%Y"))
        # str(end)

        # Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:Y1')

        ws.row_dimensions[1].height = 25

        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 35
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 12
        ws.column_dimensions['P'].width = 15

        ws.column_dimensions['Q'].width = 12
        ws.column_dimensions['R'].width = 40
        ws.column_dimensions['S'].width = 12
        ws.column_dimensions['T'].width = 12
        ws.column_dimensions['U'].width = 15
        ws.column_dimensions['V'].width = 15
        ws.column_dimensions['W'].width = 15
        ws.column_dimensions['X'].width = 15
        ws.column_dimensions['Y'].width = 15

        # Crear la cabecera
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['B3'] = 'Nº'

        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['C3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['C3'] = 'CODIGO'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['D3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['D3'] = 'COMPROBANTE'

        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['E3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['E3'] = 'TIPO'

        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['F3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['F3'] = 'ESTADO'

        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['G3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['G3'] = 'FECHA'

        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['H3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['H3'] = 'INGRESO'

        ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['I3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['I3'] = 'DOCUMENTO'

        ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['J3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['J3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['J3'] = 'PROVEEDOR'

        ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['K3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['K3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['K3'] = 'TOTAL'

        ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['L3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['L3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['L3'] = 'MONEDA'

        ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['M3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['M3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['M3'] = 'CAMBIO'

        ws['N3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['N3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['N3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['N3'] = 'SUBTOTAL(S/.)'

        ws['O3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['O3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['O3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['O3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['O3'] = 'IGV(S/.)'

        ws['P3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['P3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['P3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['P3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['P3'] = 'TOTAL(S/.)'

        ws['Q3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Q3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['Q3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['Q3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['Q3'] = 'CODIGO'

        ws['R3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['R3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['R3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['R3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['R3'] = 'PRODUCTO'

        ws['S3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['S3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['S3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['S3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['S3'] = 'MEDIDA'

        ws['T3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['T3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['T3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['T3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['T3'] = 'CANTIDAD'

        ws['U3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['U3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['U3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['U3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['U3'] = 'UNIDAD'

        ws['V3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['V3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['V3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['V3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['V3'] = 'PRECIO'

        ws['W3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['W3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['W3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['W3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['W3'] = 'TOTAL'

        ws['X3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['X3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['X3'] = 'PRECIO SOLES'

        ws['Y3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Y3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['Y3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['Y3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['Y3'] = 'TOTAL SOLES'

        total = decimal.Decimal(0.000000)
        subtotal = decimal.Decimal(0.000000)
        total_igv = decimal.Decimal(0.000000)
        for o in order_set:
            # Pintamos los datos en el reporteç
            date_one = "-"
            if o.create_at:
                date_one = o.create_at.strftime("%d-%m-%Y")
            date_two = "-"
            if o.invoice_date:
                date_two = o.invoice_date.strftime("%d-%m-%Y")
            date_three = "-"
            if o.date_document:
                date_three = o.date_document.strftime("%d-%m-%Y")
            color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
            text_color = Font(name='Calibri', size=9)
            border_cell = Border(left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 top=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))
            align_cell = Alignment(vertical="center", horizontal="center")

            # ws.cell(row=row, column=2).alignment = align_cell
            ws.cell(row=row, column=2).alignment = align_cell
            ws.cell(row=row, column=2).border = border_cell
            ws.cell(row=row, column=2).fill = color_cell
            ws.cell(row=row, column=2).font = text_color
            ws.cell(row=row, column=2).value = o.number

            ws.cell(row=row, column=3).alignment = align_cell
            ws.cell(row=row, column=3).border = border_cell
            ws.cell(row=row, column=3).fill = color_cell
            ws.cell(row=row, column=3).font = text_color
            ws.cell(row=row, column=3).value = o.get_code()

            ws.cell(row=row, column=4).alignment = align_cell
            ws.cell(row=row, column=4).border = border_cell
            ws.cell(row=row, column=4).fill = color_cell
            ws.cell(row=row, column=4).font = text_color
            ws.cell(row=row, column=4).value = o.invoice_number

            ws.cell(row=row, column=5).alignment = align_cell
            ws.cell(row=row, column=5).border = border_cell
            ws.cell(row=row, column=5).fill = color_cell
            ws.cell(row=row, column=5).font = text_color
            ws.cell(row=row, column=5).value = o.get_doc_display()

            ws.cell(row=row, column=6).alignment = align_cell
            ws.cell(row=row, column=6).border = border_cell
            ws.cell(row=row, column=6).fill = color_cell
            ws.cell(row=row, column=6).font = text_color
            ws.cell(row=row, column=6).value = o.get_status_display()

            ws.cell(row=row, column=7).alignment = align_cell
            ws.cell(row=row, column=7).border = border_cell
            ws.cell(row=row, column=7).fill = color_cell
            ws.cell(row=row, column=7).font = text_color
            ws.cell(row=row, column=7).value = date_one

            ws.cell(row=row, column=8).alignment = align_cell
            ws.cell(row=row, column=8).border = border_cell
            ws.cell(row=row, column=8).fill = color_cell
            ws.cell(row=row, column=8).font = text_color
            ws.cell(row=row, column=8).value = date_two

            ws.cell(row=row, column=9).alignment = align_cell
            ws.cell(row=row, column=9).border = border_cell
            ws.cell(row=row, column=9).fill = color_cell
            ws.cell(row=row, column=9).font = text_color
            ws.cell(row=row, column=9).value = date_three

            ws.cell(row=row, column=10).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=row, column=10).border = border_cell
            ws.cell(row=row, column=10).fill = color_cell
            ws.cell(row=row, column=10).font = text_color
            ws.cell(row=row, column=10).value = o.person.names

            ws.cell(row=row, column=11).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=row, column=11).border = border_cell
            ws.cell(row=row, column=11).fill = color_cell
            ws.cell(row=row, column=11).font = text_color
            ws.cell(row=row, column=11).value = o.total

            ws.cell(row=row, column=12).alignment = align_cell
            ws.cell(row=row, column=12).border = border_cell
            ws.cell(row=row, column=12).fill = color_cell
            ws.cell(row=row, column=12).font = text_color
            ws.cell(row=row, column=12).value = o.get_coin_display()

            ws.cell(row=row, column=13).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=row, column=13).border = border_cell
            ws.cell(row=row, column=13).fill = color_cell
            ws.cell(row=row, column=13).font = text_color
            ws.cell(row=row, column=13).value = o.change

            amount = o.change * o.total
            amount_sin_igv = amount / decimal.Decimal(1.1800)
            igv = amount - (amount / decimal.Decimal(1.1800))

            ws.cell(row=row, column=14).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=row, column=14).border = border_cell
            ws.cell(row=row, column=14).fill = color_cell
            ws.cell(row=row, column=14).font = text_color
            ws.cell(row=row, column=14).value = round(amount_sin_igv, 2)

            ws.cell(row=row, column=15).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=row, column=15).border = border_cell
            ws.cell(row=row, column=15).fill = color_cell
            ws.cell(row=row, column=15).font = text_color
            ws.cell(row=row, column=15).value = round(igv, 2)

            ws.cell(row=row, column=16).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=row, column=16).border = border_cell
            ws.cell(row=row, column=16).fill = color_cell
            ws.cell(row=row, column=16).font = text_color
            ws.cell(row=row, column=16).value = round(amount, 2)
            subtotal = subtotal + amount_sin_igv
            total_igv = total_igv + igv
            total = total + amount
            di = 0
            temp_total = 0
            for d in o.orderdetail_set.filter(is_state=True).order_by('id'):
                di += 1
                ws.cell(row=row, column=17).alignment = align_cell
                ws.cell(row=row, column=17).border = border_cell
                ws.cell(row=row, column=17).fill = color_cell
                ws.cell(row=row, column=17).font = text_color
                ws.cell(row=row, column=17).value = d.product.code

                ws.cell(row=row, column=18).alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row=row, column=18).border = border_cell
                ws.cell(row=row, column=18).fill = color_cell
                ws.cell(row=row, column=18).font = text_color
                ws.cell(row=row, column=18).value = d.product.name

                ws.cell(row=row, column=19).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=row, column=19).border = border_cell
                ws.cell(row=row, column=19).fill = color_cell
                ws.cell(row=row, column=19).font = text_color
                ws.cell(row=row, column=19).value = d.product.measure()

                ws.cell(row=row, column=20).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=row, column=20).border = border_cell
                ws.cell(row=row, column=20).fill = color_cell
                ws.cell(row=row, column=20).font = text_color
                ws.cell(row=row, column=20).value = d.quantity

                ws.cell(row=row, column=21).alignment = align_cell
                ws.cell(row=row, column=21).border = border_cell
                ws.cell(row=row, column=21).fill = color_cell
                ws.cell(row=row, column=21).font = text_color
                ws.cell(row=row, column=21).value = d.get_unit_display()

                ws.cell(row=row, column=22).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=row, column=22).border = border_cell
                ws.cell(row=row, column=22).fill = color_cell
                ws.cell(row=row, column=22).font = text_color
                ws.cell(row=row, column=22).value = d.price

                ws.cell(row=row, column=23).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=row, column=23).border = border_cell
                ws.cell(row=row, column=23).fill = color_cell
                ws.cell(row=row, column=23).font = text_color
                ws.cell(row=row, column=23).value = d.amount()

                ws.cell(row=row, column=24).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=row, column=24).border = border_cell
                ws.cell(row=row, column=24).fill = color_cell
                ws.cell(row=row, column=24).font = text_color
                ws.cell(row=row, column=24).value = d.price * o.change

                ws.cell(row=row, column=25).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=row, column=25).border = border_cell
                ws.cell(row=row, column=25).fill = color_cell
                ws.cell(row=row, column=25).font = text_color
                ws.cell(row=row, column=25).value = d.amount() * o.change

                row += 1
                temp_total += d.amount()
            # print(temp_total)
            # ws.merge_cells('B1:S1')
            cont += 1
            ws.merge_cells('B' + str(row - di) + ':B' + str(row - 1))
            # print(o.id, o.number, o.bill_serial, o.bill_number)
            ws.merge_cells('C' + str(row - di) + ':C' + str(row - 1))
            ws.merge_cells('D' + str(row - di) + ':D' + str(row - 1))
            ws.merge_cells('E' + str(row - di) + ':E' + str(row - 1))
            ws.merge_cells('F' + str(row - di) + ':F' + str(row - 1))
            ws.merge_cells('G' + str(row - di) + ':G' + str(row - 1))
            ws.merge_cells('H' + str(row - di) + ':H' + str(row - 1))
            ws.merge_cells('I' + str(row - di) + ':I' + str(row - 1))
            ws.merge_cells('J' + str(row - di) + ':J' + str(row - 1))
            ws.merge_cells('K' + str(row - di) + ':K' + str(row - 1))
            ws.merge_cells('L' + str(row - di) + ':L' + str(row - 1))
            ws.merge_cells('M' + str(row - di) + ':M' + str(row - 1))
            ws.merge_cells('N' + str(row - di) + ':N' + str(row - 1))
            ws.merge_cells('O' + str(row - di) + ':O' + str(row - 1))
            ws.merge_cells('P' + str(row - di) + ':P' + str(row - 1))

        ws.cell(row=row, column=14).alignment = Alignment(vertical="center", horizontal="right")
        ws.cell(row=row, column=14).border = Border(left=Side(border_style="thin"),
                                                    right=Side(border_style="thin"),
                                                    top=Side(border_style="thin"),
                                                    bottom=Side(border_style="thin"))
        ws.cell(row=row, column=14).fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
        ws.cell(row=row, column=14).font = Font(name='Calibri', size=10)
        ws.cell(row=row, column=14).value = round(subtotal, 2)

        ws.cell(row=row, column=15).alignment = Alignment(vertical="center", horizontal="right")
        ws.cell(row=row, column=15).border = Border(left=Side(border_style="thin"),
                                                    right=Side(border_style="thin"),
                                                    top=Side(border_style="thin"),
                                                    bottom=Side(border_style="thin"))
        ws.cell(row=row, column=15).fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
        ws.cell(row=row, column=15).font = Font(name='Calibri', size=10)
        ws.cell(row=row, column=15).value = round(total_igv, 2)

        ws.cell(row=row, column=16).alignment = Alignment(vertical="center", horizontal="right")
        ws.cell(row=row, column=16).border = Border(left=Side(border_style="thin"),
                                                    right=Side(border_style="thin"),
                                                    top=Side(border_style="thin"),
                                                    bottom=Side(border_style="thin"))
        ws.cell(row=row, column=16).fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
        ws.cell(row=row, column=16).font = Font(name='Calibri', size=10)
        ws.cell(row=row, column=16).value = round(total, 2)
        # row += 1

        # Establecer el nombre de mi archivo
        nombre_archivo = "Compras-jhunior.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


def export_product_filter(request, c1=None, c2=None, c3=None, c4=None, array_id=None):
    product_set = []
    _arr = []
    if array_id is not None and array_id != " ":
        str1 = array_id.replace(']', '').replace('[', '')
        _arr = str1.replace('"', '').replace("'", '').split(",")
    if c1 != "0" and len(c1) > 3:
        product_set = Product.objects.filter(name__contains=c1.upper()).order_by('id')
    elif c4 != "0":
        product_set = Product.objects.filter(code=int(c4))
    elif c2 != "0":
        if c3 == '1':
            product_set = Product.objects.filter(
                brand__name__contains=c2.upper()).select_related('brand', 'family').prefetch_related(
                Prefetch(
                    'presentation_set', queryset=Presentation.objects.select_related('product')
                )
            ).order_by('id')
        elif c3 == '2':
            product_set = Product.objects.filter(family__name__contains=c2.upper()).order_by('id')
        else:
            product_set = []
    elif c3 == '1':
        product_set = Product.objects.all().select_related('brand', 'family').prefetch_related(
            Prefetch(
                'presentation_set', queryset=Presentation.objects.select_related('product')
            )
        ).order_by('id')
    else:
        product_set = Product.objects.filter(id__in=_arr).order_by('id')

    wb = Workbook()
    bandera = True
    cont = 1
    row = 4
    if bandera:
        ws = wb.active
        ws.title = str("Productos Filtrados")  # hoja
        bandera = False
    else:
        ws = wb.create_sheet(str("Productos Filtrados") + str(cont))
    # Crear el título en la hoja
    ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B1'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['B1'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
    ws['B1'] = 'LISTA DE PRODUCTOS'

    # Cambiar caracteristicas de las celdas
    ws.merge_cells('B1:R1')

    ws.row_dimensions[1].height = 25

    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 9
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 9
    ws.column_dimensions['H'].width = 9
    ws.column_dimensions['I'].width = 9
    ws.column_dimensions['J'].width = 9
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 9
    ws.column_dimensions['M'].width = 9
    ws.column_dimensions['N'].width = 25
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 20
    ws.column_dimensions['Q'].width = 20
    ws.column_dimensions['R'].width = 15

    # Crear la cabecera
    ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['B3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['B3'] = 'CODIGO'

    ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['C3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['C3'] = 'NOMBRE PRODUCTO'

    ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['D3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['D3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['D3'] = 'DESCRIPCION PRODUCTO'

    ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['E3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['E3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['E3'] = 'MARCA'

    ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['F3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['F3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['F3'] = 'FAMILIA'

    ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['G3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['G3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['G3'] = 'TIPO'

    ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['H3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['H3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['H3'] = 'ANCHO'

    ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['I3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['I3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['I3'] = 'LARGO'

    ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['J3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['J3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['J3'] = 'ALTO'

    ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['K3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['K3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['K3'] = 'ALMACEN'

    ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['L3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['L3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['L3'] = 'STOCK'

    ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['M3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['M3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['M3'] = 'MINIMO'

    ws['N3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['N3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['N3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['N3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['N3'] = 'PRODUCTO RELACIONADO'

    ws['O3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['O3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['O3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['O3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['O3'] = 'CANTIDAD'

    ws['P3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['P3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['P3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['P3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['P3'] = 'UNIDAD'

    ws['Q3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['Q3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['Q3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['Q3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['Q3'] = 'CANTIDAD UNITARIA'

    ws['R3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['R3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['R3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
    ws['R3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
    ws['R3'] = 'PRICE'

    for p in product_set:
        # Pintamos los datos en el reporte
        color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
        stock_cell = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type="solid")
        text_color = Font(name='Calibri', size=9)
        border_cell = Border(left=Side(border_style="thin"),
                             right=Side(border_style="thin"),
                             top=Side(border_style="thin"),
                             bottom=Side(border_style="thin"))
        align_cell = Alignment(horizontal="center")
        ws.cell(row=row, column=2).alignment = align_cell
        ws.cell(row=row, column=2).border = border_cell
        ws.cell(row=row, column=2).fill = color_cell
        ws.cell(row=row, column=2).font = text_color
        ws.cell(row=row, column=2).value = p.code

        ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=3).border = border_cell
        ws.cell(row=row, column=3).fill = color_cell
        ws.cell(row=row, column=3).font = text_color
        ws.cell(row=row, column=3).value = p.name

        ws.cell(row=row, column=4).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=4).border = border_cell
        ws.cell(row=row, column=4).fill = color_cell
        ws.cell(row=row, column=4).font = text_color
        ws.cell(row=row, column=4).value = p.description

        ws.cell(row=row, column=5).alignment = align_cell
        ws.cell(row=row, column=5).border = border_cell
        ws.cell(row=row, column=5).fill = color_cell
        ws.cell(row=row, column=5).font = text_color
        ws.cell(row=row, column=5).value = p.brand.name

        ws.cell(row=row, column=6).alignment = align_cell
        ws.cell(row=row, column=6).border = border_cell
        ws.cell(row=row, column=6).fill = color_cell
        ws.cell(row=row, column=6).font = text_color
        ws.cell(row=row, column=6).value = p.family.name

        ws.cell(row=row, column=7).alignment = align_cell
        ws.cell(row=row, column=7).border = border_cell
        ws.cell(row=row, column=7).fill = color_cell
        ws.cell(row=row, column=7).font = text_color
        ws.cell(row=row, column=7).value = p.get_type_display()

        ws.cell(row=row, column=8).alignment = align_cell
        ws.cell(row=row, column=8).border = border_cell
        ws.cell(row=row, column=8).fill = color_cell
        ws.cell(row=row, column=8).font = text_color
        ws.cell(row=row, column=8).value = p.width

        ws.cell(row=row, column=9).alignment = align_cell
        ws.cell(row=row, column=9).border = border_cell
        ws.cell(row=row, column=9).fill = color_cell
        ws.cell(row=row, column=9).font = text_color
        ws.cell(row=row, column=9).value = p.length

        ws.cell(row=row, column=10).alignment = align_cell
        ws.cell(row=row, column=10).border = border_cell
        ws.cell(row=row, column=10).fill = color_cell
        ws.cell(row=row, column=10).font = text_color
        ws.cell(row=row, column=10).value = p.height

        ws.cell(row=row, column=11).alignment = align_cell
        ws.cell(row=row, column=11).border = border_cell
        ws.cell(row=row, column=11).fill = color_cell
        ws.cell(row=row, column=11).font = text_color
        ws.cell(row=row, column=11).value = p.store

        ws.cell(row=row, column=12).alignment = align_cell
        ws.cell(row=row, column=12).border = border_cell
        ws.cell(row=row, column=12).fill = stock_cell
        ws.cell(row=row, column=12).font = text_color
        ws.cell(row=row, column=12).value = p.stock

        ws.cell(row=row, column=13).alignment = align_cell
        ws.cell(row=row, column=13).border = border_cell
        ws.cell(row=row, column=13).fill = color_cell
        ws.cell(row=row, column=13).font = text_color
        ws.cell(row=row, column=13).value = p.minimum

        ws.cell(row=row, column=14).alignment = align_cell
        ws.cell(row=row, column=14).border = border_cell
        ws.cell(row=row, column=14).fill = color_cell
        ws.cell(row=row, column=14).font = text_color
        ws.cell(row=row, column=14).value = p.relation
        dp = p.presentation_set.count()
        df = 0

        for pt in p.presentation_set.all():
            df += 1
            ws.cell(row=row, column=15).alignment = align_cell
            ws.cell(row=row, column=15).border = border_cell
            ws.cell(row=row, column=15).fill = color_cell
            ws.cell(row=row, column=15).font = text_color
            ws.cell(row=row, column=15).value = pt.quantity

            ws.cell(row=row, column=16).alignment = align_cell
            ws.cell(row=row, column=16).border = border_cell
            ws.cell(row=row, column=16).fill = color_cell
            ws.cell(row=row, column=16).font = text_color
            ws.cell(row=row, column=16).value = pt.get_unit_display()

            ws.cell(row=row, column=17).alignment = align_cell
            ws.cell(row=row, column=17).border = border_cell
            ws.cell(row=row, column=17).fill = color_cell
            ws.cell(row=row, column=17).font = text_color
            ws.cell(row=row, column=17).value = pt.quantity_niu

            ws.cell(row=row, column=18).alignment = align_cell
            ws.cell(row=row, column=18).border = border_cell
            ws.cell(row=row, column=18).fill = color_cell
            ws.cell(row=row, column=18).font = text_color
            ws.cell(row=row, column=18).value = pt.price
            if df < dp:
                row += 1
        # ws.merge_cells('B' + str(row-df) + ':B' + str(row))
        cont += 1
        row += 1

    # Establecer el nombre de mi archivo
    nombre_archivo = "ProductFilter.xlsx"
    # Definir el tipo de respuesta que se va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


def report_kardex_cont(request, init=None, end=None, pk=None):
    if init and end and pk:
        product_obj = Product.objects.get(id=int(pk))
        # m = month[5:]
        detail_set = OrderDetail.objects.filter(product=product_obj, order__type__in=['V', 'C'],
                                                order__create_at__range=(init, end)).order_by('id')

        order_detail_obj = OrderDetail.objects.filter(product=product_obj, order__type__in=['V', 'C']).first()
        wb = Workbook()
        bandera = True
        cont = 1
        row = 4
        if bandera:
            ws = wb.active
            ws.title = str("PRODUCTO")  # hoja
            bandera = False
        else:
            ws = wb.create_sheet(str(product_obj.name) + str(cont))
        # Crear el título en la hoja
        my_date = datetime.datetime.now()
        date_now = my_date.strftime("%d-%m-%Y")
        hour_now = my_date.strftime("%H:%M:%S")
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B1'].font = Font(name='Calibri', size=12, bold=True, color='00FFFFFF')
        ws['B1'] = str(product_obj.name.upper())

        # Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:M1')

        ws.row_dimensions[1].height = 25

        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15

        # Crear la cabecera
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['B3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['B3'] = 'FECHA'

        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['C3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['C3'] = 'ORDEN'

        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['D3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['D3'] = 'OPERACIÓN'

        ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['E3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['E3'] = 'CANTIDAD'

        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['F3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['F3'] = 'MEDIDA'

        ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['G3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['G3'] = 'ENTRADA'

        ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['H3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['H3'] = 'SALIDA'

        ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['I3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['I3'] = 'SALDO'

        ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['J3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['J3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['J3'] = 'INGRESO'

        ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['K3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['K3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['K3'] = 'EGRESO'

        ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['L3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['L3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['L3'] = 'SALDO'

        ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['M3'].fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        ws['M3'].font = Font(name='Calibro', size=10, bold=True, color='00FFFFFF')
        ws['M3'] = 'PROMEDIO'
        counter_for = 0
        for d in detail_set:
            if counter_for == 0 and order_detail_obj.id == d.id:
                color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                text_color = Font(name='Calibri', size=9)
                border_cell = Border(left=Side(border_style="thin"),
                                     right=Side(border_style="thin"),
                                     top=Side(border_style="thin"),
                                     bottom=Side(border_style="thin"))
                align_cell = Alignment(horizontal="center")

                ws.cell(row=row, column=2).alignment = align_cell
                ws.cell(row=row, column=2).border = border_cell
                ws.cell(row=row, column=2).fill = color_cell
                ws.cell(row=row, column=2).font = text_color
                ws.cell(row=row, column=2).value = "-"

                ws.cell(row=row, column=3).alignment = align_cell
                ws.cell(row=row, column=3).border = border_cell
                ws.cell(row=row, column=3).fill = color_cell
                ws.cell(row=row, column=3).font = text_color
                ws.cell(row=row, column=3).value = "INICIO"

                ws.cell(row=row, column=4).alignment = align_cell
                ws.cell(row=row, column=4).border = border_cell
                ws.cell(row=row, column=4).fill = color_cell
                ws.cell(row=row, column=4).font = text_color
                ws.cell(row=row, column=4).value = "RESTANTE"

                ws.cell(row=row, column=5).alignment = align_cell
                ws.cell(row=row, column=5).border = border_cell
                ws.cell(row=row, column=5).fill = color_cell
                ws.cell(row=row, column=5).font = text_color
                ws.cell(row=row, column=5).value = "-"

                ws.cell(row=row, column=6).alignment = align_cell
                ws.cell(row=row, column=6).border = border_cell
                ws.cell(row=row, column=6).fill = color_cell
                ws.cell(row=row, column=6).font = text_color
                ws.cell(row=row, column=6).value = "UNIDAD"

                ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
                ws.cell(row=row, column=7).border = border_cell
                ws.cell(row=row, column=7).fill = color_cell
                ws.cell(row=row, column=7).font = text_color
                ws.cell(row=row, column=7).value = "-"

                ws.cell(row=row, column=8).alignment = align_cell
                ws.cell(row=row, column=8).border = border_cell
                ws.cell(row=row, column=8).fill = color_cell
                ws.cell(row=row, column=8).font = text_color
                ws.cell(row=row, column=8).value = "-"

                ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
                ws.cell(row=row, column=9).border = border_cell
                ws.cell(row=row, column=9).fill = color_cell
                ws.cell(row=row, column=9).font = text_color
                ws.cell(row=row, column=9).value = round(d.previous(), 2)

                ws.cell(row=row, column=10).alignment = Alignment(horizontal="right")
                ws.cell(row=row, column=10).border = border_cell
                ws.cell(row=row, column=10).fill = color_cell
                ws.cell(row=row, column=10).font = text_color
                ws.cell(row=row, column=10).value = "-"

                ws.cell(row=row, column=11).alignment = Alignment(horizontal="right")
                ws.cell(row=row, column=11).border = border_cell
                ws.cell(row=row, column=11).fill = color_cell
                ws.cell(row=row, column=11).font = text_color
                ws.cell(row=row, column=11).value = "-"

                ws.cell(row=row, column=12).alignment = Alignment(horizontal="right")
                ws.cell(row=row, column=12).border = border_cell
                ws.cell(row=row, column=12).fill = color_cell
                ws.cell(row=row, column=12).font = text_color
                ws.cell(row=row, column=12).value = round(d.total_initial(), 2)

                ws.cell(row=row, column=13).alignment = Alignment(horizontal="right")
                ws.cell(row=row, column=13).border = border_cell
                ws.cell(row=row, column=13).fill = color_cell
                ws.cell(row=row, column=13).font = text_color
                ws.cell(row=row, column=13).value = round(d.product.price_unit(), 2)
                row += 1
            # Estado activo
            if d.is_state:
                if d.operation == 'S':
                    color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                    text_color = Font(name='Calibri', size=9)
                    border_cell = Border(left=Side(border_style="thin"),
                                         right=Side(border_style="thin"),
                                         top=Side(border_style="thin"),
                                         bottom=Side(border_style="thin"))
                    align_cell = Alignment(horizontal="center")

                    ws.cell(row=row, column=2).alignment = align_cell
                    ws.cell(row=row, column=2).border = border_cell
                    ws.cell(row=row, column=2).fill = color_cell
                    ws.cell(row=row, column=2).font = text_color
                    ws.cell(row=row, column=2).value = d.order.create_at

                    ws.cell(row=row, column=3).alignment = align_cell
                    ws.cell(row=row, column=3).border = border_cell
                    ws.cell(row=row, column=3).fill = color_cell
                    ws.cell(row=row, column=3).font = text_color
                    ws.cell(row=row, column=3).value = str(d.order.get_type_display()).upper() + "Nº " + str(
                        d.order.number)

                    ws.cell(row=row, column=4).alignment = align_cell
                    ws.cell(row=row, column=4).border = border_cell
                    ws.cell(row=row, column=4).fill = color_cell
                    ws.cell(row=row, column=4).font = text_color
                    ws.cell(row=row, column=4).value = str(d.get_operation_display()).upper()

                    ws.cell(row=row, column=5).alignment = align_cell
                    ws.cell(row=row, column=5).border = border_cell
                    ws.cell(row=row, column=5).fill = color_cell
                    ws.cell(row=row, column=5).font = text_color
                    ws.cell(row=row, column=5).value = str(d.quantity)

                    ws.cell(row=row, column=6).alignment = align_cell
                    ws.cell(row=row, column=6).border = border_cell
                    ws.cell(row=row, column=6).fill = color_cell
                    ws.cell(row=row, column=6).font = text_color
                    ws.cell(row=row, column=6).value = str(d.get_unit_display())

                    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=7).border = border_cell
                    ws.cell(row=row, column=7).fill = color_cell
                    ws.cell(row=row, column=7).font = text_color
                    ws.cell(row=row, column=7).value = "-"

                    ws.cell(row=row, column=8).alignment = align_cell
                    ws.cell(row=row, column=8).border = border_cell
                    ws.cell(row=row, column=8).fill = color_cell
                    ws.cell(row=row, column=8).font = text_color
                    ws.cell(row=row, column=8).value = round(d.quantity_niu, 2)

                    ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=9).border = border_cell
                    ws.cell(row=row, column=9).fill = color_cell
                    ws.cell(row=row, column=9).font = text_color
                    ws.cell(row=row, column=9).value = round(d.quantity_remaining, 2)

                    ws.cell(row=row, column=10).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=10).border = border_cell
                    ws.cell(row=row, column=10).fill = color_cell
                    ws.cell(row=row, column=10).font = text_color
                    ws.cell(row=row, column=10).value = "-"

                    ws.cell(row=row, column=11).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=11).border = border_cell
                    ws.cell(row=row, column=11).fill = color_cell
                    ws.cell(row=row, column=11).font = text_color
                    ws.cell(row=row, column=11).value = d.amount()

                    ws.cell(row=row, column=12).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=12).border = border_cell
                    ws.cell(row=row, column=12).fill = color_cell
                    ws.cell(row=row, column=12).font = text_color
                    ws.cell(row=row, column=12).value = d.balance_remaining()

                    ws.cell(row=row, column=13).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=13).border = border_cell
                    ws.cell(row=row, column=13).fill = color_cell
                    ws.cell(row=row, column=13).font = text_color
                    ws.cell(row=row, column=13).value = d.price
                elif d.operation == 'E':
                    color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                    text_color = Font(name='Calibri', size=9)
                    border_cell = Border(left=Side(border_style="thin"),
                                         right=Side(border_style="thin"),
                                         top=Side(border_style="thin"),
                                         bottom=Side(border_style="thin"))
                    align_cell = Alignment(horizontal="center")

                    ws.cell(row=row, column=2).alignment = align_cell
                    ws.cell(row=row, column=2).border = border_cell
                    ws.cell(row=row, column=2).fill = color_cell
                    ws.cell(row=row, column=2).font = text_color
                    ws.cell(row=row, column=2).value = d.order.create_at

                    ws.cell(row=row, column=3).alignment = align_cell
                    ws.cell(row=row, column=3).border = border_cell
                    ws.cell(row=row, column=3).fill = color_cell
                    ws.cell(row=row, column=3).font = text_color
                    ws.cell(row=row, column=3).value = str(d.order.get_type_display()).upper() + "Nº " + str(
                        d.order.number)

                    ws.cell(row=row, column=4).alignment = align_cell
                    ws.cell(row=row, column=4).border = border_cell
                    ws.cell(row=row, column=4).fill = color_cell
                    ws.cell(row=row, column=4).font = text_color
                    ws.cell(row=row, column=4).value = str(d.get_operation_display()).upper()

                    ws.cell(row=row, column=5).alignment = align_cell
                    ws.cell(row=row, column=5).border = border_cell
                    ws.cell(row=row, column=5).fill = color_cell
                    ws.cell(row=row, column=5).font = text_color
                    ws.cell(row=row, column=5).value = str(d.quantity)

                    ws.cell(row=row, column=6).alignment = align_cell
                    ws.cell(row=row, column=6).border = border_cell
                    ws.cell(row=row, column=6).fill = color_cell
                    ws.cell(row=row, column=6).font = text_color
                    ws.cell(row=row, column=6).value = str(d.get_unit_display())

                    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=7).border = border_cell
                    ws.cell(row=row, column=7).fill = color_cell
                    ws.cell(row=row, column=7).font = text_color
                    ws.cell(row=row, column=7).value = d.quantity_niu

                    ws.cell(row=row, column=8).alignment = align_cell
                    ws.cell(row=row, column=8).border = border_cell
                    ws.cell(row=row, column=8).fill = color_cell
                    ws.cell(row=row, column=8).font = text_color
                    ws.cell(row=row, column=8).value = "-"

                    ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=9).border = border_cell
                    ws.cell(row=row, column=9).fill = color_cell
                    ws.cell(row=row, column=9).font = text_color
                    ws.cell(row=row, column=9).value = d.quantity_remaining

                    ws.cell(row=row, column=10).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=10).border = border_cell
                    ws.cell(row=row, column=10).fill = color_cell
                    ws.cell(row=row, column=10).font = text_color
                    ws.cell(row=row, column=10).value = d.amount()

                    ws.cell(row=row, column=11).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=11).border = border_cell
                    ws.cell(row=row, column=11).fill = color_cell
                    ws.cell(row=row, column=11).font = text_color
                    ws.cell(row=row, column=11).value = "-"

                    ws.cell(row=row, column=12).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=12).border = border_cell
                    ws.cell(row=row, column=12).fill = color_cell
                    ws.cell(row=row, column=12).font = text_color
                    ws.cell(row=row, column=12).value = d.balance_remaining()

                    ws.cell(row=row, column=13).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=13).border = border_cell
                    ws.cell(row=row, column=13).fill = color_cell
                    ws.cell(row=row, column=13).font = text_color
                    ws.cell(row=row, column=13).value = d.price
            else:
                if d.operation == 'S':
                    color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                    text_color = Font(name='Calibri', size=9)
                    border_cell = Border(left=Side(border_style="thin"),
                                         right=Side(border_style="thin"),
                                         top=Side(border_style="thin"),
                                         bottom=Side(border_style="thin"))
                    align_cell = Alignment(horizontal="center")

                    ws.cell(row=row, column=2).alignment = align_cell
                    ws.cell(row=row, column=2).border = border_cell
                    ws.cell(row=row, column=2).fill = color_cell
                    ws.cell(row=row, column=2).font = text_color
                    ws.cell(row=row, column=2).value = d.order.create_at

                    ws.cell(row=row, column=3).alignment = align_cell
                    ws.cell(row=row, column=3).border = border_cell
                    ws.cell(row=row, column=3).fill = color_cell
                    ws.cell(row=row, column=3).font = text_color
                    ws.cell(row=row, column=3).value = str(d.order.get_type_display()).upper() + "Nº " + str(
                        d.order.number)

                    ws.cell(row=row, column=4).alignment = align_cell
                    ws.cell(row=row, column=4).border = border_cell
                    ws.cell(row=row, column=4).fill = color_cell
                    ws.cell(row=row, column=4).font = text_color
                    ws.cell(row=row, column=4).value = str(d.get_operation_display()).upper()

                    ws.cell(row=row, column=5).alignment = align_cell
                    ws.cell(row=row, column=5).border = border_cell
                    ws.cell(row=row, column=5).fill = color_cell
                    ws.cell(row=row, column=5).font = text_color
                    ws.cell(row=row, column=5).value = str(d.quantity)

                    ws.cell(row=row, column=6).alignment = align_cell
                    ws.cell(row=row, column=6).border = border_cell
                    ws.cell(row=row, column=6).fill = color_cell
                    ws.cell(row=row, column=6).font = text_color
                    ws.cell(row=row, column=6).value = str(d.get_unit_display())

                    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=7).border = border_cell
                    ws.cell(row=row, column=7).fill = color_cell
                    ws.cell(row=row, column=7).font = text_color
                    ws.cell(row=row, column=7).value = "-"

                    ws.cell(row=row, column=8).alignment = align_cell
                    ws.cell(row=row, column=8).border = border_cell
                    ws.cell(row=row, column=8).fill = color_cell
                    ws.cell(row=row, column=8).font = text_color
                    ws.cell(row=row, column=8).value = round(d.quantity_niu, 2)

                    ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=9).border = border_cell
                    ws.cell(row=row, column=9).fill = color_cell
                    ws.cell(row=row, column=9).font = text_color
                    ws.cell(row=row, column=9).value = round(d.quantity_remaining, 2)

                    ws.cell(row=row, column=10).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=10).border = border_cell
                    ws.cell(row=row, column=10).fill = color_cell
                    ws.cell(row=row, column=10).font = text_color
                    ws.cell(row=row, column=10).value = "-"

                    ws.cell(row=row, column=11).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=11).border = border_cell
                    ws.cell(row=row, column=11).fill = color_cell
                    ws.cell(row=row, column=11).font = text_color
                    ws.cell(row=row, column=11).value = d.amount()

                    ws.cell(row=row, column=12).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=12).border = border_cell
                    ws.cell(row=row, column=12).fill = color_cell
                    ws.cell(row=row, column=12).font = text_color
                    ws.cell(row=row, column=12).value = d.balance_remaining()

                    ws.cell(row=row, column=13).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=13).border = border_cell
                    ws.cell(row=row, column=13).fill = color_cell
                    ws.cell(row=row, column=13).font = text_color
                    ws.cell(row=row, column=13).value = d.price
                elif d.operation == 'E':
                    color_cell = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
                    text_color = Font(name='Calibri', size=9)
                    border_cell = Border(left=Side(border_style="thin"),
                                         right=Side(border_style="thin"),
                                         top=Side(border_style="thin"),
                                         bottom=Side(border_style="thin"))
                    align_cell = Alignment(horizontal="center")

                    ws.cell(row=row, column=2).alignment = align_cell
                    ws.cell(row=row, column=2).border = border_cell
                    ws.cell(row=row, column=2).fill = color_cell
                    ws.cell(row=row, column=2).font = text_color
                    ws.cell(row=row, column=2).value = d.order.create_at

                    ws.cell(row=row, column=3).alignment = align_cell
                    ws.cell(row=row, column=3).border = border_cell
                    ws.cell(row=row, column=3).fill = color_cell
                    ws.cell(row=row, column=3).font = text_color
                    ws.cell(row=row, column=3).value = str(d.order.get_type_display()).upper() + "Nº " + str(
                        d.order.number)

                    ws.cell(row=row, column=4).alignment = align_cell
                    ws.cell(row=row, column=4).border = border_cell
                    ws.cell(row=row, column=4).fill = color_cell
                    ws.cell(row=row, column=4).font = text_color
                    ws.cell(row=row, column=4).value = str(d.get_operation_display()).upper()

                    ws.cell(row=row, column=5).alignment = align_cell
                    ws.cell(row=row, column=5).border = border_cell
                    ws.cell(row=row, column=5).fill = color_cell
                    ws.cell(row=row, column=5).font = text_color
                    ws.cell(row=row, column=5).value = str(d.quantity)

                    ws.cell(row=row, column=6).alignment = align_cell
                    ws.cell(row=row, column=6).border = border_cell
                    ws.cell(row=row, column=6).fill = color_cell
                    ws.cell(row=row, column=6).font = text_color
                    ws.cell(row=row, column=6).value = str(d.get_unit_display())

                    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=7).border = border_cell
                    ws.cell(row=row, column=7).fill = color_cell
                    ws.cell(row=row, column=7).font = text_color
                    ws.cell(row=row, column=7).value = d.quantity_niu

                    ws.cell(row=row, column=8).alignment = align_cell
                    ws.cell(row=row, column=8).border = border_cell
                    ws.cell(row=row, column=8).fill = color_cell
                    ws.cell(row=row, column=8).font = text_color
                    ws.cell(row=row, column=8).value = "-"

                    ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=9).border = border_cell
                    ws.cell(row=row, column=9).fill = color_cell
                    ws.cell(row=row, column=9).font = text_color
                    ws.cell(row=row, column=9).value = d.quantity_remaining

                    ws.cell(row=row, column=10).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=10).border = border_cell
                    ws.cell(row=row, column=10).fill = color_cell
                    ws.cell(row=row, column=10).font = text_color
                    ws.cell(row=row, column=10).value = d.amount()

                    ws.cell(row=row, column=11).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=11).border = border_cell
                    ws.cell(row=row, column=11).fill = color_cell
                    ws.cell(row=row, column=11).font = text_color
                    ws.cell(row=row, column=11).value = "-"

                    ws.cell(row=row, column=12).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=12).border = border_cell
                    ws.cell(row=row, column=12).fill = color_cell
                    ws.cell(row=row, column=12).font = text_color
                    ws.cell(row=row, column=12).value = d.balance_remaining()

                    ws.cell(row=row, column=13).alignment = Alignment(horizontal="right")
                    ws.cell(row=row, column=13).border = border_cell
                    ws.cell(row=row, column=13).fill = color_cell
                    ws.cell(row=row, column=13).font = text_color
                    ws.cell(row=row, column=13).value = d.price
            counter_for += 1
            cont += 1
            row += 1

        # Establecer el nombre de mi archivo
        nombre_archivo = "Kardex.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


def truncate_sheet_name(name, max_length=31):
    return name[:max_length]


def sanitize_sheet_name(sheet_name):
    invalid_chars = '[]:*?/\\'
    for char in invalid_chars:
        sheet_name = sheet_name.replace(char, '-')
    return sheet_name[:31]


def get_unique_sheet_name(sheet_name, workbook):
    original_name = sheet_name
    counter = 1
    # Asegúrate de que el nombre base no exceda 29 caracteres para permitir el sufijo
    base_name = original_name[:29]

    while sheet_name.lower() in (s.name.lower() for s in workbook.worksheets()):
        sheet_name = f"{base_name}-{counter}"  # Deja espacio para el sufijo numérico
        counter += 1
        # Asegúrate de que el nuevo nombre no exceda 31 caracteres
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
    return sheet_name


def kardex_excel(request, month_year=None):
    # last_kardex = Kardex.objects.filter(product=OuterRef('id')).order_by('-id')

    year, month = map(int, month_year.split('-'))
    last_month = month - 1 if month > 1 else 12

    last_kardex = Kardex.objects.filter(
        product=OuterRef('id'),
        create_at__month=last_month
    ).order_by('-create_at', '-id')

    alt_last_kardex = Kardex.objects.filter(
        product=OuterRef('id'),
        operation='C',
        create_at__month=month
    )

    first_day_of_month = datetime.date(year, last_month, 1)

    fallback_last_kardex = Kardex.objects.filter(
        product=OuterRef('id'),
        create_at__lt=first_day_of_month  # Filtra cualquier fecha antes del mes actual
    ).order_by('-create_at')

    current_month_purchases = Kardex.objects.filter(
        product=OuterRef('id'),
        operation='E',
        type_operation='02',
        create_at__month=month
    ).values('product').annotate(total=Sum('quantity')).values('total')

    current_month_sales = Kardex.objects.filter(
        product=OuterRef('id'),
        operation='S',
        type_operation='01',
        create_at__month=month
    ).values('product').annotate(total=Sum('quantity')).values('total')

    last_month_remaining_quantity = Kardex.objects.filter(
        product=OuterRef('id'),
        create_at__month=last_month
    ).order_by('-create_at', '-id').values('remaining_quantity')[:1]

    current_purchases_total_price_total = Kardex.objects.filter(
        product=OuterRef('id'),
        operation='E',
        type_operation='02',
        create_at__month=month
    ).values('product').annotate(total_price_total=Sum('price_total')).values('total_price_total')

    current_sales_total_price_total = Kardex.objects.filter(
        product=OuterRef('id'),
        operation='S',
        type_operation='01',
        create_at__month=month
    ).values('product').annotate(total_price_total=Sum('price_total')).values('total_price_total')

    current_remaining_price = Kardex.objects.filter(
        product=OuterRef('id'),
        operation='E',
        type_operation='02',
        create_at__month=month
    ).order_by('-create_at', '-id').values('remaining_price')[:1]

    last_month_remaining_price_total = Kardex.objects.filter(
        product=OuterRef('id'),
        create_at__month=last_month
    ).order_by('-create_at', '-id').values('remaining_price_total')[:1]

    initial_remaining_quantity = Kardex.objects.filter(
        product=OuterRef('id'),
        operation='C',
        type_operation='16',
        create_at__month=month,
        order_detail__order__type='C'
    ).values('remaining_quantity')[:1]

    initial_remaining_price_total = Kardex.objects.filter(
        product=OuterRef('id'),
        operation='C',
        type_operation='16',
        create_at__month=month,
        order_detail__order__type='C'
    ).values('remaining_price_total')[:1]

    products_set = Product.objects.annotate(
        last_kardex_id=Subquery(last_kardex.values('id')[:1]),
        last_kardex_operation=Subquery(last_kardex.values('operation')[:1]),
        last_kardex_quantity=Subquery(last_kardex.values('quantity')[:1]),
        last_kardex_remaining_quantity=Subquery(last_kardex.values('remaining_quantity')[:1]),
        last_kardex_remaining_price=Subquery(last_kardex.values('remaining_price')[:1]),
        last_kardex_remaining_price_total=Subquery(last_kardex.values('remaining_price_total')[:1]),

        alt_last_kardex_id=Subquery(alt_last_kardex.values('id')[:1]),
        alt_last_kardex_operation=Subquery(alt_last_kardex.values('operation')[:1]),
        alt_last_kardex_quantity=Subquery(alt_last_kardex.values('quantity')[:1]),
        alt_last_kardex_remaining_quantity=Subquery(alt_last_kardex.values('remaining_quantity')[:1]),
        alt_last_kardex_remaining_price=Subquery(alt_last_kardex.values('remaining_price')[:1]),
        alt_last_kardex_remaining_price_total=Subquery(alt_last_kardex.values('remaining_price_total')[:1]),
        alt_last_kardex_order_detail=Subquery(alt_last_kardex.values('order_detail')[:1]),
        alt_last_kardex_type_order=Subquery(alt_last_kardex.values('order_detail__order__type')[:1]),

        fallback_last_kardex_kardex_id=Subquery(fallback_last_kardex.values('id')[:1]),
        fallback_last_kardex_kardex_operation=Subquery(fallback_last_kardex.values('operation')[:1]),
        fallback_last_kardex_kardex_quantity=Subquery(fallback_last_kardex.values('quantity')[:1]),
        fallback_last_kardex_kardex_remaining_quantity=Subquery(fallback_last_kardex.values('remaining_quantity')[:1]),
        fallback_last_kardex_kardex_remaining_price=Subquery(fallback_last_kardex.values('remaining_price')[:1]),
        fallback_last_kardex_kardex_remaining_price_total=Subquery(fallback_last_kardex.values('remaining_price_total')[:1]),


        current_month_purchases=Coalesce(Subquery(current_month_purchases),
                                         Value(0, output_field=DecimalField(max_digits=10, decimal_places=6))),
        current_month_sales=Coalesce(Subquery(current_month_sales),
                                     Value(0, output_field=DecimalField(max_digits=10, decimal_places=6))),
        current_remaining_price=Subquery(current_remaining_price.values('remaining_price')[:1]),

        last_month_quantity_init=Coalesce(Subquery(last_month_remaining_quantity),
                                          Value(0, output_field=DecimalField(max_digits=10, decimal_places=6))),

        new_remaining_quantity=(F('last_month_quantity_init') + F('current_month_purchases')) - F('current_month_sales'),
        current_purchases_total_price_total=Coalesce(Subquery(current_purchases_total_price_total),
                                                     Value(0, output_field=DecimalField(max_digits=30, decimal_places=15))),
        current_sales_total_price_total=Coalesce(Subquery(current_sales_total_price_total),
                                                 Value(0, output_field=DecimalField(max_digits=30, decimal_places=15))),
        last_month_remaining_price_total=Coalesce(Subquery(last_month_remaining_price_total),
                                                  Value(0, output_field=DecimalField(max_digits=30, decimal_places=15))),
        new_remaining_price_total=(F('last_month_remaining_price_total') + F('current_purchases_total_price_total')) - F(
            'current_sales_total_price_total'),
        initial_remaining_quantity=Coalesce(Subquery(initial_remaining_quantity),
                                            Value(0, output_field=DecimalField(max_digits=10, decimal_places=6))),
        initial_remaining_price_total=Coalesce(Subquery(initial_remaining_price_total),
                                               Value(0, output_field=DecimalField(max_digits=10, decimal_places=6))),

    ).exclude(last_kardex_remaining_quantity=None,alt_last_kardex_remaining_quantity=None,
              fallback_last_kardex_kardex_remaining_quantity=None)
        # .filter(
        # last_kardex_quantity__isnull=False,
        # id=1)
    # ).exclude(last_kardex_remaining_quantity=0, last_kardex_remaining_price=0,
    #           last_kardex_remaining_price_total=0).order_by('name')

    # for p in products_set:
    #     print(p.id)
    #     print("last_kardex_id", p.last_kardex_id)
    #     print("last_kardex_operation", p.last_kardex_operation)
    #     print("last_kardex_quantity", p.last_kardex_quantity)
    #     print("last_kardex_remaining_quantity", p.last_kardex_remaining_quantity)
    #     print("last_kardex_remaining_price", p.last_kardex_remaining_price)
    #     print("last_kardex_remaining_price_total", p.last_kardex_remaining_price_total)
    #     print("--------------------------------")
    #
    #     print("alt_last_kardex_id", p.alt_last_kardex_id)
    #     print("alt_last_kardex_operation", p.alt_last_kardex_operation)
    #     print("alt_last_kardex_quantity", p.alt_last_kardex_quantity)
    #     print("alt_last_kardex_remaining_quantity", p.alt_last_kardex_remaining_quantity)
    #     print("alt_last_kardex_remaining_price", p.alt_last_kardex_remaining_price)
    #     print("alt_last_kardex_remaining_price_total", p.alt_last_kardex_remaining_price_total)
    #     print("--------------------------------")
    #
    #     print("fallback_last_kardex_kardex_id", p.fallback_last_kardex_kardex_id)
    #     print("fallback_last_kardex_kardex_operation", p.fallback_last_kardex_kardex_operation)
    #     print("fallback_last_kardex_kardex_quantity", p.fallback_last_kardex_kardex_quantity)
    #     print("fallback_last_kardex_kardex_remaining_quantity", p.fallback_last_kardex_kardex_remaining_quantity)
    #     print("fallback_last_kardex_kardex_remaining_price", p.fallback_last_kardex_kardex_remaining_price)
    #     print("fallback_last_kardex_kardex_remaining_price_total", p.fallback_last_kardex_kardex_remaining_price_total)
    #     print("--------------------------------")
    #
    #
    #     print("current_month_purchases", p.current_month_purchases)
    #     print("current_month_sales", p.current_month_sales)
    #     print("last_month_quantity_init", p.last_month_quantity_init)
    #     print("current_purchases_total_price_total", p.current_purchases_total_price_total)
    #     print("current_sales_total_price_total", p.current_sales_total_price_total)
    #     print("current_remaining_price", p.current_remaining_price)
    #     print("new_remaining_quantity", p.new_remaining_quantity)
    #     print("new_remaining_price_total", p.new_remaining_price_total)
    #     print("initial_remaining_quantity", p.initial_remaining_quantity)
    #     print("initial_remaining_price_total", p.initial_remaining_price_total)
    # #
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=REPORTE_DE_KARDEX.xlsx'
    #
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        workbook = writer.book

        header_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'font_name': 'Arial',
            'font_size': 9
        })

        worksheet = workbook.add_worksheet('Kardex General')

        last_day = calendar.monthrange(year, month)[1]
        last_date = f"{year}-{month:02d}-{last_day:02d}"

        worksheet.write('A1', 'Empresa', workbook.add_format({'align': 'left', 'valign': 'vcenter', 'font_name': 'Arial','font_size': 9}))
        worksheet.write('A2', 'Reporte:', workbook.add_format({'align': 'left', 'valign': 'vcenter', 'font_name': 'Arial','font_size': 9}))
        worksheet.write('A3', 'KARDEX AL {}'.format(last_date), workbook.add_format({'valign': 'vcenter', 'font_name': 'Arial','font_size': 9}))
        worksheet.write('B1', '20600854535 - PERNOS JHUNIOR S.R.L.', workbook.add_format({'align': 'left','valign': 'vbot', 'font_name': 'Arial','font_size': 9}))
        worksheet.write('B2', 'Reporte de Analisis de Stock', workbook.add_format({'align': 'left', 'valign': 'vbot', 'font_name': 'Arial','font_size': 9}))

        worksheet.set_row(0, 12.75)  # Fila 1
        worksheet.set_row(1, 12.75)  # Fila 2
        worksheet.set_row(2, 12.75)  # Fila 3
        worksheet.set_row(3, 12.75)  # Fila 4

        headers = [
            'Nº', 'Codigo', 'Detalle', 'SaldoIF', 'EntradaF', 'SalidaF', 'SaldoF',
            'SaldoI', 'EntradaS', 'SalidaS', 'SaldoS', 'CostoProm'
        ]

        worksheet.merge_range(2, 3, 2, 6, 'FISICO', header_format)
        worksheet.merge_range(2, 7, 2, 10, 'VALORES', header_format)

        worksheet.write_row(3, 0, headers, header_format)
        worksheet.set_row(3, 36.75)

        # Formatos de columnas
        worksheet.set_column('A:A', 7.57)  # Nº
        worksheet.set_column('B:B', 7.3)  # Codigo
        worksheet.set_column('C:C', 58.86)  # Detalle
        worksheet.set_column('D:D', 11.43)  # SaldoIF
        worksheet.set_column('E:E', 11.43)  # EntradaF
        worksheet.set_column('F:F', 11.43)  # SalidaF
        worksheet.set_column('G:G', 11.43)  # SaldoF
        worksheet.set_column('H:H', 11.43)  # SaldoI
        worksheet.set_column('I:I', 11.43)  # EntradaS
        worksheet.set_column('J:J', 11.43)  # SalidaS
        worksheet.set_column('K:K', 11.43)  # SaldoS
        worksheet.set_column('L:L', 13)  # CostoProm

        worksheet.freeze_panes(4, 0)

        numeric_format = workbook.add_format({
            'num_format': '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-',
            'align': 'right',
            'font_name': 'Arial',
            'font_size': 9,
            'border': 1
        })
        numeric_format_remaining = workbook.add_format({
            'num_format': '_ * #,##0.000000_ ;_ * -#,##0.000000_ ;_ * "-"??_ ;_ @_ ',
            'align': 'right',
            'font_name': 'Arial',
            'font_size': 9,
            'border': 1
        })
        numeric_format_bg = workbook.add_format({
            'num_format': '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-',
            'align': 'right',
            'font_name': 'Arial',
            'font_size': 9,
            'border': 1,
            'bg_color': '#C6E0B4',
        })
        cell_format = workbook.add_format({
            'font_name': 'Arial',
            'font_size': 9,
            'align': 'left',
            'border': 1
        })
        cell_format_1 = workbook.add_format({
            'font_name': 'Arial',
            'font_size': 9,
            'align': 'center',
            'border': 1
        })
        cell_format_2 = workbook.add_format({
            'font_name': 'Arial',
            'font_size': 9,
            'align': 'right',
            'border': 1
        })

        for row_num, product in enumerate(products_set, start=1):
            worksheet.set_row(row_num + 3, 12.75)  # Altura de fila de 12.75
            worksheet.write(row_num + 3, 0, row_num, cell_format_1)  # Nº
            worksheet.write(row_num + 3, 1, int(product.code), cell_format_2)  # Codigo
            worksheet.write(row_num + 3, 2, product.name.strip() + ' ' + product.measure(), cell_format)  # Detalle

            if product.last_kardex_remaining_quantity is not None:
                last_month_quantity_init = product.last_kardex_remaining_quantity
            elif product.alt_last_kardex_remaining_quantity is not None and product.alt_last_kardex_remaining_quantity > 0 and product.alt_last_kardex_order_detail is None:
                last_month_quantity_init = product.alt_last_kardex_remaining_quantity
            elif product.alt_last_kardex_remaining_quantity is not None and product.alt_last_kardex_remaining_quantity > 0 and product.alt_last_kardex_order_detail is not None and product.alt_last_kardex_type_order == 'C':
                last_month_quantity_init = 0
            elif product.fallback_last_kardex_kardex_remaining_quantity is not None and product.fallback_last_kardex_kardex_remaining_quantity > 0:
                last_month_quantity_init = product.fallback_last_kardex_kardex_remaining_quantity
            else:
                last_month_quantity_init = 0
            worksheet.write(row_num + 3, 3, last_month_quantity_init, numeric_format)  # SaldoIF

            # entry_F_value = (
            #     product.initial_remaining_quantity
            #     if product.last_kardex_operation == 'C' and product.current_month_purchases == 0
            #     else product.current_month_purchases
            # )
            if product.last_kardex_operation == 'C' and product.current_month_purchases == 0:
                entry_F_value = product.initial_remaining_quantity
            elif product.alt_last_kardex_remaining_quantity is not None and product.alt_last_kardex_remaining_quantity > 0 and product.alt_last_kardex_order_detail is not None and product.alt_last_kardex_type_order == 'C':
                entry_F_value = product.alt_last_kardex_remaining_quantity
            else:
                entry_F_value = product.current_month_purchases

            worksheet.write(row_num + 3, 4, entry_F_value, numeric_format)  # EntradaF

            # if product.last_kardex_operation == 'C':
            #     worksheet.write(row_num + 3, 4, product.initial_remaining_quantity, numeric_format)  # EntradaF
            # else:
            #     if product.current_month_purchases != 0:
            #         worksheet.write(row_num + 3, 4, product.current_month_purchases, numeric_format)  # EntradaF
            #     else:
            #         worksheet.write(row_num + 3, 4, product.initial_remaining_quantity, numeric_format)  # EntradaF

            worksheet.write(row_num + 3, 5, product.current_month_sales, numeric_format)  # SalidaF

            new_remaining_quantity = (decimal.Decimal(last_month_quantity_init) + decimal.Decimal(entry_F_value)) - decimal.Decimal(product.current_month_sales)
            # if product.new_remaining_quantity != 0:
            #     worksheet.write(row_num + 3, 6, product.new_remaining_quantity, numeric_format_bg)  # SaldoF
            # else:
            #     worksheet.write(row_num + 3, 6, product.initial_remaining_quantity, numeric_format_bg)  # SaldoF

            worksheet.write(row_num + 3, 6, new_remaining_quantity, numeric_format_bg)  # SaldoF

            if product.last_kardex_remaining_price_total is not None:
                last_kardex_remaining_price_total = product.last_kardex_remaining_price_total
            elif product.alt_last_kardex_remaining_price_total is not None and product.alt_last_kardex_remaining_price_total > 0 and product.alt_last_kardex_order_detail is None:
                last_kardex_remaining_price_total = product.alt_last_kardex_remaining_price_total
            elif product.alt_last_kardex_remaining_price_total is not None and product.alt_last_kardex_remaining_price_total > 0 and product.alt_last_kardex_order_detail is not None and product.alt_last_kardex_type_order == 'C':
                last_kardex_remaining_price_total = 0
            elif product.fallback_last_kardex_kardex_remaining_price_total is not None and product.fallback_last_kardex_kardex_remaining_price_total > 0:
                last_kardex_remaining_price_total = product.fallback_last_kardex_kardex_remaining_price_total
            else:
                last_kardex_remaining_price_total = 0

            worksheet.write(row_num + 3, 7, last_kardex_remaining_price_total, numeric_format)  # SaldoI

            # entry_S_value = (
            #     product.initial_remaining_price_total
            #     if product.last_kardex_operation == 'C' and product.current_purchases_total_price_total == 0
            #     else product.current_purchases_total_price_total
            # )

            if product.last_kardex_operation == 'C' and product.current_purchases_total_price_total == 0:
                entry_S_value = product.initial_remaining_price_total
            elif product.alt_last_kardex_remaining_price_total is not None and product.alt_last_kardex_remaining_price_total > 0 and product.alt_last_kardex_order_detail is not None and product.alt_last_kardex_type_order == 'C':
                entry_S_value = product.alt_last_kardex_remaining_price_total
            else:
                entry_S_value = product.current_purchases_total_price_total

            worksheet.write(row_num + 3, 8, entry_S_value, numeric_format)  # EntradaS

            # if product.last_kardex_operation == 'C':
            #     worksheet.write(row_num + 3, 8, product.initial_remaining_price_total, numeric_format)  # EntradaS
            # else:
            #     if product.current_purchases_total_price_total != 0:
            #         worksheet.write(row_num + 3, 8, product.current_purchases_total_price_total, numeric_format)  # EntradaS
            #     else:
            #         worksheet.write(row_num + 3, 8, product.initial_remaining_price_total, numeric_format)  # EntradaS

            worksheet.write(row_num + 3, 9, product.current_sales_total_price_total, numeric_format)  # SalidaS
            # if product.new_remaining_price_total != 0:
            #     worksheet.write(row_num + 3, 10, product.new_remaining_price_total, numeric_format_bg)  # SaldoS
            # else:
            #     worksheet.write(row_num + 3, 10, product.initial_remaining_price_total, numeric_format_bg)  # SaldoS

            new_remaining_price_total = (decimal.Decimal(last_kardex_remaining_price_total) + decimal.Decimal(entry_S_value)) - decimal.Decimal(product.current_sales_total_price_total)

            worksheet.write(row_num + 3, 10, new_remaining_price_total, numeric_format_bg)  # SaldoS

            # if product.last_kardex_remaining_price is not None:
            #     last_kardex_remaining_price = product.last_kardex_remaining_price
            # elif product.alt_last_kardex_remaining_price is not None and product.alt_last_kardex_remaining_price > 0 and product.alt_last_kardex_order_detail is None:
            #     last_kardex_remaining_price = product.alt_last_kardex_remaining_price
            # elif product.alt_last_kardex_remaining_price is not None and product.alt_last_kardex_remaining_price > 0 and product.alt_last_kardex_order_detail is not None and product.alt_last_kardex_type_order == 'C':
            #     last_kardex_remaining_price = 0
            # elif product.fallback_last_kardex_kardex_remaining_price is not None and product.fallback_last_kardex_kardex_remaining_price > 0:
            #     last_kardex_remaining_price = product.fallback_last_kardex_kardex_remaining_price
            # else:
            #     last_kardex_remaining_price = 0

            if product.fallback_last_kardex_kardex_remaining_price is not None:
                last_kardex_remaining_price = product.fallback_last_kardex_kardex_remaining_price
            elif product.alt_last_kardex_remaining_price is not None and product.alt_last_kardex_remaining_price > 0:
                last_kardex_remaining_price = product.alt_last_kardex_remaining_price
            elif product.current_remaining_price is not None:
                last_kardex_remaining_price = product.current_remaining_price
            elif product.last_kardex_remaining_price is not None and product.last_kardex_remaining_price > 0:
                last_kardex_remaining_price = product.last_kardex_remaining_price
            else:
                last_kardex_remaining_price = 0

            worksheet.write(row_num + 3, 11, round(last_kardex_remaining_price, 6), numeric_format_remaining)  # CostoProm

    return response

